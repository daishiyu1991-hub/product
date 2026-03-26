#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
飞书集成 SDK — 亚马逊精益产品开发 Pipeline
==========================================
纯 Python + requests 实现，零外部依赖。
支持 Webhook 群通知、App API 消息、飞书文档、多维表格、文件上传。

CLI 用法:
  python feishu_sdk.py webhook --text "Phase 1 完成"
  python feishu_sdk.py card --template phase1 --data '{"product":"steamer","decision":"GO"}'
  python feishu_sdk.py notify --phase 1 --product steamer --decision GO --metrics '{...}'
  python feishu_sdk.py doc --md ./report.md --title "Phase 1 报告"
  python feishu_sdk.py bitable --xlsx ./data.xlsx --title "选品数据"
  python feishu_sdk.py upload --file ./dashboard.html --folder fldcnXXXX
  python feishu_sdk.py test
"""

import argparse
import base64
import hashlib
import hmac
import json
import os
import re
import sys
import time
from pathlib import Path

try:
    import requests
except ImportError:
    print("错误: 需要 requests 库。运行: pip install requests")
    sys.exit(1)


# ============================================================================
# 常量
# ============================================================================

FEISHU_BASE_URL = "https://open.feishu.cn/open-apis"
LARK_BASE_URL = "https://open.larksuite.com/open-apis"  # 国际版

PROXY_ADDR = "http://127.0.0.1:7897"

# 颜色映射（飞书卡片模板颜色）
CARD_COLORS = {
    "green": "green",
    "red": "red",
    "yellow": "yellow",
    "orange": "orange",
    "blue": "blue",
    "purple": "purple",
    "grey": "grey",
}


# ============================================================================
# 工具函数
# ============================================================================

def _get_env(key: str, default: str = "") -> str:
    """从环境变量获取配置"""
    return os.environ.get(key, default)


def _get_base_url() -> str:
    """根据环境变量决定使用国内还是国际版 API"""
    if _get_env("FEISHU_USE_LARK", "").lower() in ("1", "true", "yes"):
        return LARK_BASE_URL
    return FEISHU_BASE_URL


def _make_request(method: str, url: str, headers: dict = None,
                  json_data: dict = None, data: dict = None,
                  files: dict = None, timeout: int = 30,
                  use_proxy: bool = True) -> dict:
    """
    统一 HTTP 请求，自动处理代理和重试。
    策略：先直连，失败后走代理。
    """
    proxies = None
    if use_proxy and _get_env("FEISHU_PROXY"):
        proxies = {"https": _get_env("FEISHU_PROXY"), "http": _get_env("FEISHU_PROXY")}
    elif use_proxy:
        # 默认代理
        proxies = {"https": PROXY_ADDR, "http": PROXY_ADDR}

    kwargs = {
        "headers": headers,
        "timeout": timeout,
    }
    if json_data is not None:
        kwargs["json"] = json_data
    if data is not None:
        kwargs["data"] = data
    if files is not None:
        kwargs["files"] = files

    # 先尝试直连
    try:
        resp = getattr(requests, method)(url, **kwargs)
        return resp.json() if resp.content else {}
    except (requests.exceptions.ConnectionError, requests.exceptions.Timeout):
        pass

    # 直连失败，走代理
    if proxies:
        try:
            kwargs["proxies"] = proxies
            resp = getattr(requests, method)(url, **kwargs)
            return resp.json() if resp.content else {}
        except Exception as e:
            print(f"请求失败 (含代理): {e}")
            return {"code": -1, "msg": str(e)}

    return {"code": -1, "msg": "连接失败，无可用代理"}


# ============================================================================
# FeishuAuth - 认证管理
# ============================================================================

class FeishuAuth:
    """
    飞书认证管理器
    - Webhook 模式：无需认证，直接用 URL
    - App 模式：自动获取和缓存 tenant_access_token（有效期 2 小时）
    """

    def __init__(self, app_id: str = None, app_secret: str = None):
        self.app_id = app_id or _get_env("FEISHU_APP_ID")
        self.app_secret = app_secret or _get_env("FEISHU_APP_SECRET")
        self.base_url = _get_base_url()
        self._token = None
        self._token_expire = 0

    @property
    def has_app_credentials(self) -> bool:
        return bool(self.app_id and self.app_secret)

    def get_token(self) -> str:
        """获取 tenant_access_token，自动缓存和续期"""
        if not self.has_app_credentials:
            raise ValueError(
                "未配置 App 凭据。请设置 FEISHU_APP_ID 和 FEISHU_APP_SECRET 环境变量。\n"
                "或运行 /feishu-setup 进行配置。"
            )

        # 缓存未过期则复用（提前 5 分钟续期）
        if self._token and time.time() < self._token_expire - 300:
            return self._token

        url = f"{self.base_url}/auth/v3/tenant_access_token/internal"
        data = _make_request("post", url, json_data={
            "app_id": self.app_id,
            "app_secret": self.app_secret,
        }, use_proxy=True)

        if data.get("code") != 0:
            raise Exception(f"认证失败: {data.get('msg', data)}")

        self._token = data["tenant_access_token"]
        self._token_expire = time.time() + data.get("expire", 7200)
        return self._token

    def headers(self) -> dict:
        """返回带 Authorization 的请求头"""
        return {
            "Authorization": f"Bearer {self.get_token()}",
            "Content-Type": "application/json; charset=utf-8",
        }


# ============================================================================
# FeishuMessenger - 消息发送
# ============================================================================

class FeishuMessenger:
    """
    飞书消息发送器
    - Webhook：群机器人推送（无需 App）
    - App API：个人/群消息，支持富文本和卡片
    """

    def __init__(self, auth: FeishuAuth = None):
        self.auth = auth or FeishuAuth()
        self.base_url = _get_base_url()

    # ---- Webhook 方式 ----

    @staticmethod
    def _gen_webhook_sign(timestamp: str, secret: str) -> str:
        """生成 Webhook 签名"""
        string_to_sign = f"{timestamp}\n{secret}"
        hmac_code = hmac.new(
            string_to_sign.encode("utf-8"), digestmod=hashlib.sha256
        ).digest()
        return base64.b64encode(hmac_code).decode("utf-8")

    def send_webhook_text(self, text: str, webhook_url: str = None,
                          secret: str = None) -> dict:
        """通过 Webhook 发送纯文本消息到群聊"""
        url = webhook_url or _get_env("FEISHU_WEBHOOK_URL")
        if not url:
            raise ValueError("未配置 Webhook URL。请设置 FEISHU_WEBHOOK_URL。")

        secret = secret or _get_env("FEISHU_WEBHOOK_SECRET")
        payload = {
            "msg_type": "text",
            "content": {"text": text}
        }

        if secret:
            timestamp = str(int(time.time()))
            payload["timestamp"] = timestamp
            payload["sign"] = self._gen_webhook_sign(timestamp, secret)

        return _make_request("post", url, json_data=payload)

    def send_webhook_card(self, card: dict, webhook_url: str = None,
                          secret: str = None) -> dict:
        """通过 Webhook 发送交互式卡片到群聊"""
        url = webhook_url or _get_env("FEISHU_WEBHOOK_URL")
        if not url:
            raise ValueError("未配置 Webhook URL。请设置 FEISHU_WEBHOOK_URL。")

        secret = secret or _get_env("FEISHU_WEBHOOK_SECRET")
        payload = {
            "msg_type": "interactive",
            "card": card
        }

        if secret:
            timestamp = str(int(time.time()))
            payload["timestamp"] = timestamp
            payload["sign"] = self._gen_webhook_sign(timestamp, secret)

        return _make_request("post", url, json_data=payload)

    def send_webhook_rich_text(self, title: str, content_lines: list,
                               webhook_url: str = None, secret: str = None) -> dict:
        """通过 Webhook 发送富文本消息"""
        url = webhook_url or _get_env("FEISHU_WEBHOOK_URL")
        if not url:
            raise ValueError("未配置 Webhook URL。")

        secret = secret or _get_env("FEISHU_WEBHOOK_SECRET")
        payload = {
            "msg_type": "post",
            "content": {
                "post": {
                    "zh_cn": {
                        "title": title,
                        "content": content_lines
                    }
                }
            }
        }

        if secret:
            timestamp = str(int(time.time()))
            payload["timestamp"] = timestamp
            payload["sign"] = self._gen_webhook_sign(timestamp, secret)

        return _make_request("post", url, json_data=payload)

    # ---- App API 方式 ----

    def send_to_chat(self, chat_id: str, msg_type: str,
                     content: dict) -> dict:
        """通过 App API 向群聊发消息"""
        url = f"{self.base_url}/im/v1/messages"
        return _make_request("post", url,
            headers=self.auth.headers(),
            json_data={
                "receive_id": chat_id,
                "msg_type": msg_type,
                "content": json.dumps(content, ensure_ascii=False),
            },
            use_proxy=True
        )

    def send_to_user(self, user_id: str, msg_type: str,
                     content: dict, id_type: str = "open_id") -> dict:
        """通过 App API 向个人发消息"""
        url = f"{self.base_url}/im/v1/messages"
        params = f"?receive_id_type={id_type}"
        return _make_request("post", url + params,
            headers=self.auth.headers(),
            json_data={
                "receive_id": user_id,
                "msg_type": msg_type,
                "content": json.dumps(content, ensure_ascii=False),
            },
            use_proxy=True
        )

    def send_card_to_chat(self, chat_id: str, card: dict) -> dict:
        """向群聊发送卡片消息（App API）"""
        return self.send_to_chat(chat_id, "interactive", card)

    def send_card_to_user(self, user_id: str, card: dict) -> dict:
        """向个人发送卡片消息（App API）"""
        return self.send_to_user(user_id, "interactive", card)

    # ---- 智能发送（自动选通道） ----

    def smart_send_text(self, text: str, target: str = None) -> dict:
        """
        智能发送文本消息：
        1. 有 Webhook → 走 Webhook
        2. 有 App + chat_id → 走 App API
        3. 都没有 → 报错
        """
        webhook_url = _get_env("FEISHU_WEBHOOK_URL")
        if webhook_url:
            return self.send_webhook_text(text)

        if self.auth.has_app_credentials:
            chat_id = target or _get_env("FEISHU_DEFAULT_CHAT_ID")
            if chat_id:
                return self.send_to_chat(chat_id, "text", {"text": text})

        raise ValueError("未配置任何飞书发送通道。请先运行 /feishu-setup。")

    def smart_send_card(self, card: dict, target: str = None) -> dict:
        """智能发送卡片消息"""
        webhook_url = _get_env("FEISHU_WEBHOOK_URL")
        if webhook_url:
            return self.send_webhook_card(card)

        if self.auth.has_app_credentials:
            chat_id = target or _get_env("FEISHU_DEFAULT_CHAT_ID")
            if chat_id:
                return self.send_card_to_chat(chat_id, card)

        raise ValueError("未配置任何飞书发送通道。请先运行 /feishu-setup。")


# ============================================================================
# FeishuDocBuilder - 飞书文档创建
# ============================================================================

class FeishuDocBuilder:
    """将 Markdown 内容转换为飞书文档"""

    def __init__(self, auth: FeishuAuth = None):
        self.auth = auth or FeishuAuth()
        self.base_url = _get_base_url()

    def create_doc(self, title: str, folder_token: str = None) -> dict:
        """创建空白飞书文档，返回 {document_id, url}"""
        folder = folder_token or _get_env("FEISHU_DEFAULT_FOLDER_TOKEN")
        url = f"{self.base_url}/docx/v1/documents"
        payload = {"title": title}
        if folder:
            payload["folder_token"] = folder

        result = _make_request("post", url,
            headers=self.auth.headers(),
            json_data=payload
        )

        if result.get("code") != 0:
            raise Exception(f"创建文档失败: {result.get('msg', result)}")

        doc_data = result.get("data", {}).get("document", {})
        return {
            "document_id": doc_data.get("document_id"),
            "title": doc_data.get("title"),
            "url": doc_data.get("url", ""),
        }

    def create_block(self, document_id: str, block_id: str,
                     children: list, index: int = -1) -> dict:
        """在文档指定位置创建内容块"""
        url = f"{self.base_url}/docx/v1/documents/{document_id}/blocks/{block_id}/children"
        payload = {"children": children}
        if index >= 0:
            payload["index"] = index

        return _make_request("post", url,
            headers=self.auth.headers(),
            json_data=payload
        )

    def md_to_blocks(self, md_content: str) -> list:
        """
        将 Markdown 转换为飞书文档 Block 列表
        支持：标题、段落、粗体、列表、代码块、分割线、表格（简化版）
        """
        blocks = []
        lines = md_content.split("\n")
        i = 0
        in_code_block = False
        code_lines = []
        code_lang = ""

        while i < len(lines):
            line = lines[i]

            # 代码块处理
            if line.strip().startswith("```"):
                if in_code_block:
                    # 结束代码块
                    blocks.append(self._code_block("\n".join(code_lines), code_lang))
                    code_lines = []
                    code_lang = ""
                    in_code_block = False
                else:
                    # 开始代码块
                    in_code_block = True
                    code_lang = line.strip()[3:].strip()
                i += 1
                continue

            if in_code_block:
                code_lines.append(line)
                i += 1
                continue

            # 空行跳过
            if not line.strip():
                i += 1
                continue

            # 分割线
            if re.match(r'^-{3,}$|^\*{3,}$|^_{3,}$', line.strip()):
                blocks.append({"block_type": 22})  # divider
                i += 1
                continue

            # 标题
            heading_match = re.match(r'^(#{1,9})\s+(.+)$', line)
            if heading_match:
                level = len(heading_match.group(1))
                text = heading_match.group(2).strip()
                # 飞书 heading block_type: 3=h1, 4=h2, ..., 11=h9
                blocks.append({
                    "block_type": 2 + level,
                    f"heading{level}": {
                        "elements": self._parse_inline(text)
                    }
                })
                i += 1
                continue

            # 无序列表
            if re.match(r'^[\-\*]\s+', line):
                text = re.sub(r'^[\-\*]\s+', '', line)
                blocks.append({
                    "block_type": 15,  # bullet
                    "bullet": {
                        "elements": self._parse_inline(text)
                    }
                })
                i += 1
                continue

            # 有序列表
            ordered_match = re.match(r'^\d+\.\s+(.+)$', line)
            if ordered_match:
                text = ordered_match.group(1)
                blocks.append({
                    "block_type": 16,  # ordered
                    "ordered": {
                        "elements": self._parse_inline(text)
                    }
                })
                i += 1
                continue

            # 引用块
            if line.startswith("> "):
                text = line[2:]
                blocks.append({
                    "block_type": 14,  # quote
                    "quote": {
                        "elements": self._parse_inline(text)
                    }
                })
                i += 1
                continue

            # 普通段落
            blocks.append({
                "block_type": 2,  # text
                "text": {
                    "elements": self._parse_inline(line)
                }
            })
            i += 1

        return blocks

    def _parse_inline(self, text: str) -> list:
        """解析行内格式：**bold**, *italic*, `code`, [link](url)"""
        elements = []
        # 简化处理：按段拆分
        parts = re.split(r'(\*\*[^*]+\*\*|\*[^*]+\*|`[^`]+`|\[[^\]]+\]\([^)]+\))', text)

        for part in parts:
            if not part:
                continue

            # 粗体
            bold_match = re.match(r'^\*\*(.+)\*\*$', part)
            if bold_match:
                elements.append({
                    "text_run": {
                        "content": bold_match.group(1),
                        "text_element_style": {"bold": True}
                    }
                })
                continue

            # 斜体
            italic_match = re.match(r'^\*(.+)\*$', part)
            if italic_match:
                elements.append({
                    "text_run": {
                        "content": italic_match.group(1),
                        "text_element_style": {"italic": True}
                    }
                })
                continue

            # 行内代码
            code_match = re.match(r'^`(.+)`$', part)
            if code_match:
                elements.append({
                    "text_run": {
                        "content": code_match.group(1),
                        "text_element_style": {"inline_code": True}
                    }
                })
                continue

            # 链接
            link_match = re.match(r'^\[(.+)\]\((.+)\)$', part)
            if link_match:
                elements.append({
                    "text_run": {
                        "content": link_match.group(1),
                        "text_element_style": {
                            "link": {"url": link_match.group(2)}
                        }
                    }
                })
                continue

            # 普通文本
            elements.append({
                "text_run": {"content": part}
            })

        return elements

    def _code_block(self, code: str, language: str = "") -> dict:
        """生成代码块"""
        lang_map = {
            "python": 49, "javascript": 19, "json": 20,
            "bash": 3, "shell": 3, "html": 15, "css": 9,
            "": 1,  # PlainText
        }
        return {
            "block_type": 18,  # code
            "code": {
                "elements": [{"text_run": {"content": code}}],
                "language": lang_map.get(language.lower(), 1),
            }
        }

    def create_doc_from_md(self, md_path: str, title: str = None,
                           folder_token: str = None) -> dict:
        """
        一键将本地 MD 文件创建为飞书文档
        返回 {document_id, title, url}
        """
        md_path = Path(md_path)
        if not md_path.exists():
            raise FileNotFoundError(f"文件不存在: {md_path}")

        title = title or md_path.stem
        md_content = md_path.read_text(encoding="utf-8")

        # 1. 创建空文档
        doc_info = self.create_doc(title, folder_token)
        doc_id = doc_info["document_id"]

        # 2. 转换 MD → blocks
        blocks = self.md_to_blocks(md_content)

        if blocks:
            # 3. 获取文档根 block_id
            get_url = f"{self.base_url}/docx/v1/documents/{doc_id}/blocks/{doc_id}"
            root = _make_request("get", get_url, headers=self.auth.headers())
            root_block_id = doc_id  # 根 block 的 id 就是 document_id

            # 4. 分批写入（每批 50 个 block）
            batch_size = 50
            for j in range(0, len(blocks), batch_size):
                batch = blocks[j:j + batch_size]
                self.create_block(doc_id, root_block_id, batch)

        print(f"✅ 飞书文档已创建: {doc_info.get('url', doc_id)}")
        return doc_info


# ============================================================================
# FeishuBitableWriter - 多维表格写入
# ============================================================================

class FeishuBitableWriter:
    """将 XLSX/CSV 数据写入飞书多维表格"""

    FIELD_TYPE_MAP = {
        "str": 1,       # 文本
        "int": 2,       # 数字
        "float": 2,     # 数字
        "bool": 7,      # 复选框
        "date": 5,      # 日期
        "datetime": 5,  # 日期时间
    }

    def __init__(self, auth: FeishuAuth = None):
        self.auth = auth or FeishuAuth()
        self.base_url = _get_base_url()

    def create_bitable_app(self, name: str, folder_token: str = None) -> str:
        """创建多维表格应用，返回 app_token"""
        folder = folder_token or _get_env("FEISHU_DEFAULT_FOLDER_TOKEN")
        url = f"{self.base_url}/bitable/v1/apps"
        payload = {"name": name}
        if folder:
            payload["folder_token"] = folder

        result = _make_request("post", url,
            headers=self.auth.headers(),
            json_data=payload
        )

        if result.get("code") != 0:
            raise Exception(f"创建多维表格失败: {result.get('msg', result)}")

        return result["data"]["app"]["app_token"]

    def create_table(self, app_token: str, name: str, fields: list) -> str:
        """创建数据表，返回 table_id"""
        url = f"{self.base_url}/bitable/v1/apps/{app_token}/tables"
        payload = {
            "table": {
                "name": name,
                "default_view_name": "默认视图",
                "fields": fields,
            }
        }

        result = _make_request("post", url,
            headers=self.auth.headers(),
            json_data=payload
        )

        if result.get("code") != 0:
            raise Exception(f"创建数据表失败: {result.get('msg', result)}")

        return result["data"]["table_id"]

    def batch_create_records(self, app_token: str, table_id: str,
                             records: list) -> dict:
        """批量创建记录（每批最多 500 条）"""
        url = f"{self.base_url}/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_create"

        # 转换为飞书格式
        feishu_records = [{"fields": record} for record in records]

        return _make_request("post", url,
            headers=self.auth.headers(),
            json_data={"records": feishu_records}
        )

    def _infer_fields(self, headers: list, rows: list) -> list:
        """根据数据自动推断字段类型"""
        fields = []
        for h in headers:
            if h is None:
                continue
            sample_values = [r.get(str(h)) for r in rows[:20] if r.get(str(h)) is not None]

            if not sample_values:
                field_type = 1  # 默认文本
            elif all(isinstance(v, bool) for v in sample_values):
                field_type = 7  # 复选框
            elif all(isinstance(v, (int, float)) for v in sample_values):
                field_type = 2  # 数字
            else:
                field_type = 1  # 文本

            fields.append({
                "field_name": str(h),
                "type": field_type,
            })

        return fields

    def _chunk(self, lst: list, size: int):
        """分批生成器"""
        for i in range(0, len(lst), size):
            yield lst[i:i + size]

    def xlsx_to_bitable(self, xlsx_path: str, title: str = None,
                        folder_token: str = None) -> dict:
        """
        读取 XLSX → 自动创建多维表格 → 写入全部数据
        返回 {app_token, tables: [{table_id, name, records}]}
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("需要 openpyxl 库。运行: pip install openpyxl")

        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            raise FileNotFoundError(f"文件不存在: {xlsx_path}")

        title = title or xlsx_path.stem
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

        # 创建多维表格
        app_token = self.create_bitable_app(title, folder_token)
        print(f"📊 多维表格已创建: {app_token}")

        tables_result = []
        for ws in wb.worksheets:
            # 提取表头
            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if not header_row or not header_row[0]:
                continue
            headers = [str(h) if h is not None else f"列{i}"
                      for i, h in enumerate(header_row[0])]

            # 提取数据
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                record = {}
                for j, val in enumerate(row):
                    if j < len(headers) and val is not None:
                        # 转换非序列化类型
                        if hasattr(val, 'isoformat'):
                            val = val.isoformat()
                        record[headers[j]] = val
                if record:
                    rows.append(record)

            if not rows:
                continue

            # 推断字段类型并创建表
            fields = self._infer_fields(headers, rows)
            table_id = self.create_table(app_token, ws.title or "Sheet1", fields)

            # 批量写入
            total = 0
            for batch in self._chunk(rows, 500):
                self.batch_create_records(app_token, table_id, batch)
                total += len(batch)
                print(f"  ↳ {ws.title}: 已写入 {total}/{len(rows)} 条")

            tables_result.append({
                "table_id": table_id,
                "name": ws.title,
                "records": len(rows),
            })

        result = {"app_token": app_token, "tables": tables_result}
        print(f"✅ 多维表格写入完成: {len(tables_result)} 张表")
        return result

    def csv_to_bitable(self, csv_path: str, title: str = None,
                       folder_token: str = None) -> dict:
        """读取 CSV → 多维表格"""
        import csv as csv_mod

        csv_path = Path(csv_path)
        if not csv_path.exists():
            raise FileNotFoundError(f"文件不存在: {csv_path}")

        title = title or csv_path.stem

        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv_mod.DictReader(f)
            headers = reader.fieldnames or []
            rows = list(reader)

        if not rows:
            raise ValueError(f"CSV 文件为空: {csv_path}")

        app_token = self.create_bitable_app(title, folder_token)
        fields = self._infer_fields(headers, rows)
        table_id = self.create_table(app_token, csv_path.stem, fields)

        total = 0
        for batch in self._chunk(rows, 500):
            self.batch_create_records(app_token, table_id, batch)
            total += len(batch)

        print(f"✅ CSV → 多维表格完成: {total} 条记录")
        return {"app_token": app_token, "table_id": table_id, "records": total}


# ============================================================================
# FeishuUploader - 文件上传
# ============================================================================

class FeishuUploader:
    """上传文件到飞书云空间"""

    def __init__(self, auth: FeishuAuth = None):
        self.auth = auth or FeishuAuth()
        self.base_url = _get_base_url()

    def upload_file(self, file_path: str, folder_token: str = None,
                    file_name: str = None) -> dict:
        """
        上传文件到指定文件夹
        返回 {file_token, url}
        """
        folder = folder_token or _get_env("FEISHU_DEFAULT_FOLDER_TOKEN")
        if not folder:
            raise ValueError("未指定目标文件夹。请设置 FEISHU_DEFAULT_FOLDER_TOKEN。")

        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        file_name = file_name or file_path.name
        file_size = file_path.stat().st_size

        url = f"{self.base_url}/drive/v1/files/upload_all"
        headers = {"Authorization": f"Bearer {self.auth.get_token()}"}

        with open(file_path, 'rb') as f:
            result = _make_request("post", url,
                headers=headers,
                data={
                    "file_name": file_name,
                    "parent_type": "explorer",
                    "parent_node": folder,
                    "size": str(file_size),
                },
                files={"file": (file_name, f)},
            )

        if result.get("code") != 0:
            raise Exception(f"上传失败: {result.get('msg', result)}")

        file_token = result.get("data", {}).get("file_token", "")
        print(f"✅ 文件已上传: {file_name} → {file_token}")
        return {"file_token": file_token, "file_name": file_name}


# ============================================================================
# 统一入口函数：notify_feishu()
# ============================================================================

def notify_feishu(phase: str, product: str, decision: str,
                  metrics: dict = None, report_path: str = None,
                  xlsx_path: str = None, push_doc: bool = False,
                  push_bitable: bool = False, target: str = None) -> dict:
    """
    统一推送入口 — 各 Phase Skill 调用此函数

    1. 始终发送决策卡片通知
    2. 可选创建飞书文档（push_doc=True）
    3. 可选写入多维表格（push_bitable=True）
    """
    # 延迟导入卡片模板
    from card_templates import get_phase_card

    results = {"card_sent": False, "doc_created": False, "bitable_created": False}

    # 1. 发送卡片通知
    try:
        card = get_phase_card(phase, product, decision, metrics or {})
        messenger = FeishuMessenger()
        resp = messenger.smart_send_card(card, target)
        results["card_sent"] = True
        results["card_response"] = resp
        print(f"✅ 卡片消息已发送 (Phase {phase} - {decision})")
    except Exception as e:
        print(f"⚠️ 卡片发送失败: {e}")
        # 降级为纯文本
        try:
            text = f"📊 Phase {phase} 完成\n产品: {product}\n决策: {decision}"
            messenger = FeishuMessenger()
            messenger.smart_send_text(text, target)
            results["card_sent"] = True
            print("✅ 已降级为纯文本消息")
        except Exception as e2:
            print(f"❌ 消息发送失败: {e2}")

    # 2. 可选：创建飞书文档
    if push_doc and report_path:
        try:
            auth = FeishuAuth()
            doc_builder = FeishuDocBuilder(auth)
            doc_info = doc_builder.create_doc_from_md(
                report_path,
                f"Phase {phase} - {product}"
            )
            results["doc_created"] = True
            results["doc_info"] = doc_info
        except Exception as e:
            print(f"⚠️ 文档创建失败: {e}")

    # 3. 可选：写入多维表格
    if push_bitable and xlsx_path:
        try:
            auth = FeishuAuth()
            bitable_writer = FeishuBitableWriter(auth)
            bt_info = bitable_writer.xlsx_to_bitable(
                xlsx_path,
                f"Phase {phase} Data - {product}"
            )
            results["bitable_created"] = True
            results["bitable_info"] = bt_info
        except Exception as e:
            print(f"⚠️ 多维表格创建失败: {e}")

    return results


# ============================================================================
# CLI 接口
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="飞书集成 SDK — 亚马逊精益产品开发",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    subparsers = parser.add_subparsers(dest="command", help="子命令")

    # ---- webhook ----
    p_webhook = subparsers.add_parser("webhook", help="通过 Webhook 发送文本消息")
    p_webhook.add_argument("--text", required=True, help="消息文本")
    p_webhook.add_argument("--url", help="Webhook URL（覆盖环境变量）")

    # ---- card ----
    p_card = subparsers.add_parser("card", help="发送卡片消息")
    p_card.add_argument("--template", required=True,
                       help="卡片模板名: phase1/phase3/phase6/phase8/generic")
    p_card.add_argument("--data", required=True,
                       help="JSON 格式数据")

    # ---- notify ----
    p_notify = subparsers.add_parser("notify", help="统一推送（卡片+可选文档/表格）")
    p_notify.add_argument("--phase", required=True, help="Phase 编号 (1-8)")
    p_notify.add_argument("--product", required=True, help="产品名称")
    p_notify.add_argument("--decision", required=True, help="决策结果")
    p_notify.add_argument("--metrics", default="{}", help="JSON 格式指标数据")
    p_notify.add_argument("--report", help="MD 报告路径")
    p_notify.add_argument("--xlsx", help="XLSX 数据路径")
    p_notify.add_argument("--push-doc", action="store_true", help="同时创建飞书文档")
    p_notify.add_argument("--push-bitable", action="store_true",
                         help="同时写入多维表格")

    # ---- doc ----
    p_doc = subparsers.add_parser("doc", help="将 MD 文件创建为飞书文档")
    p_doc.add_argument("--md", required=True, help="Markdown 文件路径")
    p_doc.add_argument("--title", help="文档标题（默认用文件名）")
    p_doc.add_argument("--folder", help="目标文件夹 token")

    # ---- bitable ----
    p_bt = subparsers.add_parser("bitable", help="将 XLSX/CSV 写入多维表格")
    p_bt.add_argument("--xlsx", help="XLSX 文件路径")
    p_bt.add_argument("--csv", help="CSV 文件路径")
    p_bt.add_argument("--title", help="表格标题（默认用文件名）")
    p_bt.add_argument("--folder", help="目标文件夹 token")

    # ---- upload ----
    p_upload = subparsers.add_parser("upload", help="上传文件到飞书云空间")
    p_upload.add_argument("--file", required=True, help="文件路径")
    p_upload.add_argument("--folder", help="目标文件夹 token")
    p_upload.add_argument("--name", help="文件名（默认用原文件名）")

    # ---- test ----
    subparsers.add_parser("test", help="测试飞书连通性")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return

    # ---- 执行命令 ----

    if args.command == "webhook":
        messenger = FeishuMessenger()
        result = messenger.send_webhook_text(args.text, args.url)
        print(f"发送结果: {json.dumps(result, ensure_ascii=False, indent=2)}")

    elif args.command == "card":
        from card_templates import get_phase_card
        data = json.loads(args.data)
        template = args.template.replace("phase", "")
        card = get_phase_card(
            phase=template,
            product=data.get("product", ""),
            decision=data.get("decision", ""),
            metrics=data.get("metrics", data),
        )
        messenger = FeishuMessenger()
        result = messenger.smart_send_card(card)
        print(f"发送结果: {json.dumps(result, ensure_ascii=False, indent=2)}")

    elif args.command == "notify":
        metrics = json.loads(args.metrics) if args.metrics else {}
        result = notify_feishu(
            phase=args.phase,
            product=args.product,
            decision=args.decision,
            metrics=metrics,
            report_path=args.report,
            xlsx_path=args.xlsx,
            push_doc=args.push_doc,
            push_bitable=args.push_bitable,
        )
        print(f"\n推送完成: {json.dumps(result, ensure_ascii=False, indent=2, default=str)}")

    elif args.command == "doc":
        auth = FeishuAuth()
        builder = FeishuDocBuilder(auth)
        result = builder.create_doc_from_md(args.md, args.title, args.folder)
        print(f"结果: {json.dumps(result, ensure_ascii=False, indent=2)}")

    elif args.command == "bitable":
        auth = FeishuAuth()
        writer = FeishuBitableWriter(auth)
        if args.xlsx:
            result = writer.xlsx_to_bitable(args.xlsx, args.title, args.folder)
        elif args.csv:
            result = writer.csv_to_bitable(args.csv, args.title, args.folder)
        else:
            print("错误: 请指定 --xlsx 或 --csv")
            return
        print(f"结果: {json.dumps(result, ensure_ascii=False, indent=2, default=str)}")

    elif args.command == "upload":
        auth = FeishuAuth()
        uploader = FeishuUploader(auth)
        result = uploader.upload_file(args.file, args.folder, args.name)
        print(f"结果: {json.dumps(result, ensure_ascii=False, indent=2)}")

    elif args.command == "test":
        # 导入并运行测试
        test_connection()


def test_connection():
    """测试飞书连通性"""
    print("=" * 60)
    print("飞书集成连通性测试")
    print("=" * 60)

    results = {"webhook": None, "app_auth": None, "proxy": None}

    # 1. 检查 Webhook
    webhook_url = _get_env("FEISHU_WEBHOOK_URL")
    if webhook_url:
        print(f"\n✅ Webhook URL 已配置: {webhook_url[:50]}...")
        try:
            messenger = FeishuMessenger()
            resp = messenger.send_webhook_text("🔧 飞书集成测试消息 — 如果你看到这条消息，说明配置成功！")
            if resp.get("code") == 0 or resp.get("StatusCode") == 0:
                print("   ✅ Webhook 消息发送成功！")
                results["webhook"] = True
            else:
                print(f"   ❌ Webhook 消息发送失败: {resp}")
                results["webhook"] = False
        except Exception as e:
            print(f"   ❌ Webhook 异常: {e}")
            results["webhook"] = False
    else:
        print("\n⚠️ Webhook URL 未配置 (FEISHU_WEBHOOK_URL)")
        results["webhook"] = None

    # 2. 检查 App 凭据
    app_id = _get_env("FEISHU_APP_ID")
    app_secret = _get_env("FEISHU_APP_SECRET")
    if app_id and app_secret:
        print(f"\n✅ App 凭据已配置: {app_id}")
        try:
            auth = FeishuAuth()
            token = auth.get_token()
            print(f"   ✅ Token 获取成功: {token[:20]}...")
            results["app_auth"] = True
        except Exception as e:
            print(f"   ❌ 认证失败: {e}")
            results["app_auth"] = False
    else:
        print("\n⚠️ App 凭据未配置 (FEISHU_APP_ID / FEISHU_APP_SECRET)")
        results["app_auth"] = None

    # 3. 检查代理
    print(f"\n🔌 代理配置: {_get_env('FEISHU_PROXY', PROXY_ADDR)}")
    try:
        resp = requests.get("https://open.feishu.cn",
                          proxies={"https": PROXY_ADDR},
                          timeout=5)
        print(f"   ✅ 代理连接正常 (HTTP {resp.status_code})")
        results["proxy"] = True
    except Exception:
        try:
            resp = requests.get("https://open.feishu.cn", timeout=5)
            print(f"   ✅ 直连正常 (HTTP {resp.status_code})，无需代理")
            results["proxy"] = True
        except Exception as e:
            print(f"   ⚠️ 连接飞书 API 失败: {e}")
            results["proxy"] = False

    # 4. 检查其他配置
    print(f"\n📋 其他配置:")
    print(f"   默认群聊 ID: {_get_env('FEISHU_DEFAULT_CHAT_ID') or '未配置'}")
    print(f"   默认用户 ID: {_get_env('FEISHU_DEFAULT_USER_ID') or '未配置'}")
    print(f"   默认文件夹:  {_get_env('FEISHU_DEFAULT_FOLDER_TOKEN') or '未配置'}")

    # 汇总
    print("\n" + "=" * 60)
    configured = sum(1 for v in results.values() if v is True)
    total = sum(1 for v in results.values() if v is not None)
    if total == 0:
        print("⚠️ 尚未配置任何飞书凭据。请运行 /feishu-setup 开始配置。")
    else:
        print(f"测试完成: {configured}/{total} 项通过")
        if configured == total:
            print("🎉 所有配置项测试通过！")
    print("=" * 60)

    return results


if __name__ == "__main__":
    main()
