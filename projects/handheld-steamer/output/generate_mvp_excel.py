import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ========================================
# 样式定义
# ========================================
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
subheader_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
subheader_font = Font(name='Arial', bold=True, size=10)
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
orange_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
title_font = Font(name='Arial', bold=True, size=14, color='2F5496')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def apply_border(ws, min_row, max_row, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).font = normal_font

# ========================================
# Sheet 1: 单件成本分析
# ========================================
ws1 = wb.active
ws1.title = '单件成本分析'
ws1.sheet_properties.tabColor = '2F5496'

ws1.merge_cells('A1:E1')
ws1['A1'] = 'Handheld Steam Cleaner — MVP 单件成本分析'
ws1['A1'].font = title_font

headers = ['成本项', '金额 (USD)', '占售价比例', '说明', '数据来源']
for col, h in enumerate(headers, 1):
    ws1.cell(row=3, column=col, value=h)
apply_header_style(ws1, 3, 5)

data = [
    ['采购成本 (FOB)', 13.90, '', '1688 ¥82 + 定制升级 ¥18 = ¥100, 汇率 7.2', '1688 + 推测'],
    ['头程运费 (海运)', 2.50, '', '约 $150/CBM, 单件 ~0.012 CBM', '推测'],
    ['FBA 配送费', 5.80, '', '标准尺寸 小号超大, ~1.5kg', '推测'],
    ['FBA 仓储费', 0.50, '', '月均仓储 (非旺季)', '推测'],
    ['亚马逊佣金 (15%)', 6.00, '', '15% × $39.99', 'Amazon 费率'],
    ['广告分摊', 4.00, '', 'TACOS 10% 估算', '推测'],
    ['其他费用', 1.50, '', '退货损耗+包装+标签+保险', '推测'],
    ['总成本', 34.20, '', '', ''],
    ['', '', '', '', ''],
    ['售价', 39.99, '', '', ''],
    ['单件毛利', 5.79, '', '', ''],
    ['毛利率', '14.5%', '', '', ''],
]

for i, row in enumerate(data, 4):
    for j, val in enumerate(row):
        cell = ws1.cell(row=i, column=j+1, value=val)
        if j == 1 and isinstance(val, (int, float)):
            cell.number_format = '$#,##0.00'
    if row[0] in ['总成本', '单件毛利', '毛利率']:
        for j in range(5):
            ws1.cell(row=i, column=j+1).font = bold_font
            ws1.cell(row=i, column=j+1).fill = subheader_fill

# 计算占售价比例
for i in range(4, 11):
    val = ws1.cell(row=i, column=2).value
    if isinstance(val, (int, float)):
        ws1.cell(row=i, column=3, value=val/39.99)
        ws1.cell(row=i, column=3).number_format = '0.0%'

apply_border(ws1, 3, 15, 5)
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 45
ws1.column_dimensions['E'].width = 18

# ========================================
# Sheet 2: 首批投入预算
# ========================================
ws2 = wb.create_sheet('首批投入预算')
ws2.sheet_properties.tabColor = '548235'

ws2.merge_cells('A1:D1')
ws2['A1'] = '首批投入预算明细 (500件)'
ws2['A1'].font = title_font

headers = ['投入项', '金额 (USD)', '说明', '数据来源']
for col, h in enumerate(headers, 1):
    ws2.cell(row=3, column=col, value=h)
apply_header_style(ws2, 3, 4)

data = [
    ['首批采购 (500件)', 6950, '$13.90 × 500', '1688 + 推测'],
    ['模具/定制费', 3000, '密封系统+配件锁定结构开模', '推测'],
    ['头程运费', 1250, '$2.50 × 500', '推测'],
    ['UL/ETL 认证', 5000, '电器安全认证 ($3,000-8,000)', '推测'],
    ['产品摄影 + A+', 2000, '7张主图 + A+ 内容制作', '推测'],
    ['首月广告预算', 3000, '日均 $100, 覆盖首月', '推测'],
    ['FBA 入仓操作', 500, '贴标+入仓', '推测'],
    ['备用金', 3000, '应急/补单/促销缓冲', '推测'],
    ['总投入', 24700, '', ''],
]

for i, row in enumerate(data, 4):
    for j, val in enumerate(row):
        cell = ws2.cell(row=i, column=j+1, value=val)
        if j == 1 and isinstance(val, (int, float)):
            cell.number_format = '$#,##0'
    if row[0] == '总投入':
        for j in range(4):
            ws2.cell(row=i, column=j+1).font = bold_font
            ws2.cell(row=i, column=j+1).fill = subheader_fill

apply_border(ws2, 3, 12, 4)
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 18

# ========================================
# Sheet 3: 盈亏平衡模型
# ========================================
ws3 = wb.create_sheet('盈亏平衡模型')
ws3.sheet_properties.tabColor = 'C00000'

ws3.merge_cells('A1:G1')
ws3['A1'] = '盈亏平衡 & 12个月财务预测'
ws3['A1'].font = title_font

# 场景对比
ws3['A3'] = '场景对比分析'
ws3['A3'].font = bold_font

headers = ['场景', '售价', '日均单量', '月销量', '单件毛利', '月毛利', '回本周期']
for col, h in enumerate(headers, 1):
    ws3.cell(row=4, column=col, value=h)
apply_header_style(ws3, 4, 7)

scenarios = [
    ['保守', 39.99, 10, 300, 5.79, 1737, '14个月'],
    ['基准', 42.99, 15, 450, 8.79, 3956, '6个月'],
    ['乐观', 44.99, 20, 600, 10.79, 6474, '4个月'],
]

fills = [red_fill, yellow_fill, green_fill]
for i, (row, fill) in enumerate(zip(scenarios, fills), 5):
    for j, val in enumerate(row):
        cell = ws3.cell(row=i, column=j+1, value=val)
        cell.fill = fill
        if j in [1, 4]:
            cell.number_format = '$#,##0.00'
        elif j == 5:
            cell.number_format = '$#,##0'

apply_border(ws3, 4, 7, 7)

# 12个月预测 (基准场景)
ws3['A10'] = '12个月财务预测 (基准场景: $42.99, 日均15单)'
ws3['A10'].font = bold_font

headers2 = ['月份', '售价', '日均单量', '月销量', '月销额', '月总成本', '月毛利', '累计毛利', '累计投资回报']
for col, h in enumerate(headers2, 1):
    ws3.cell(row=11, column=col, value=h)
apply_header_style(ws3, 11, 9)

monthly_data = [
    [1, 34.99, 8, 240, 8398, 8208, 190, -24510, '-'],
    [2, 37.99, 12, 360, 13676, 12312, 1364, -23146, '-'],
    [3, 39.99, 15, 450, 17996, 15390, 2606, -20540, '-'],
    [4, 42.99, 15, 450, 19346, 15390, 3956, -16584, '-'],
    [5, 42.99, 18, 540, 23215, 18468, 4747, -11837, '-'],
    [6, 42.99, 20, 600, 25794, 20520, 5274, -6563, '-'],
    [7, 44.99, 20, 600, 26994, 20520, 6474, -89, '-'],
    [8, 44.99, 22, 660, 29693, 22572, 7121, 7032, '28%'],
    [9, 44.99, 22, 660, 29693, 22572, 7121, 14153, '57%'],
    [10, 44.99, 25, 750, 33743, 25650, 8093, 22246, '90%'],
    [11, 44.99, 28, 840, 37792, 28728, 9064, 31310, '127%'],
    [12, 44.99, 30, 900, 40491, 30780, 9711, 41021, '166%'],
]

for i, row in enumerate(monthly_data, 12):
    for j, val in enumerate(row):
        cell = ws3.cell(row=i, column=j+1, value=val)
        if j in [1]:
            cell.number_format = '$#,##0.00'
        elif j in [4, 5, 6, 7]:
            cell.number_format = '$#,##0'
        if j == 7:
            if isinstance(val, (int, float)) and val >= 0:
                cell.fill = green_fill
            elif isinstance(val, (int, float)):
                cell.fill = red_fill

apply_border(ws3, 11, 23, 9)

for col in range(1, 10):
    ws3.column_dimensions[get_column_letter(col)].width = 15

# ========================================
# Sheet 4: 功能矩阵
# ========================================
ws4 = wb.create_sheet('功能矩阵')
ws4.sheet_properties.tabColor = 'ED7D31'

ws4.merge_cells('A1:E1')
ws4['A1'] = 'MVP 功能矩阵 (Must-have / Nice-to-have / Cut)'
ws4['A1'].font = title_font

# Must-have
ws4['A3'] = '🟢 Must-have (必须做)'
ws4['A3'].font = Font(name='Arial', bold=True, size=12, color='548235')

headers = ['#', '功能', '规格要求', '决策依据', '数据来源']
for col, h in enumerate(headers, 1):
    ws4.cell(row=4, column=col, value=h)
apply_header_style(ws4, 4, 5)

must_have = [
    ['M1', '无化学剂清洁', '纯水蒸汽清洁', '100%覆盖率, 98.7%销量', 'Sorftime'],
    ['M2', '多表面适用', '地板/瓷砖/玻璃/沙发/汽车', '94.1%覆盖率', 'Sorftime'],
    ['M3', '快速加热', '≤3分钟出蒸汽', '64.7%覆盖率', 'Sorftime'],
    ['M4', '防漏水设计', '双密封圈+旋锁安全盖', '#1竞品痛点', 'CURMEDI差评'],
    ['M5', '配件稳固连接', '卡扣锁定机构', '#3竞品痛点', 'CURMEDI差评'],
    ['M6', '安全保护', '防干烧+安全泄压阀', '#4竞品痛点', 'CURMEDI差评'],
    ['M7', '多功能配件套装', '≥15件配件', '47.1%覆盖率', 'Sorftime'],
    ['M8', '手持式设计', '≤1.2kg空机', '41.2%覆盖率', 'Sorftime'],
    ['M9', 'UL/ETL认证', '电器安全认证', '准入门槛', '推测'],
]

for i, row in enumerate(must_have, 5):
    for j, val in enumerate(row):
        cell = ws4.cell(row=i, column=j+1, value=val)
        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

apply_border(ws4, 4, 13, 5)

# Nice-to-have
ws4['A16'] = '🟡 Nice-to-have (尝试加入)'
ws4['A16'].font = Font(name='Arial', bold=True, size=12, color='BF8F00')

for col, h in enumerate(headers, 1):
    ws4.cell(row=17, column=col, value=h)
apply_header_style(ws4, 17, 5)

nice_to_have = [
    ['N1', '透明水位窗口', '水箱侧面透明窗', '#8痛点:无法看水位', '+¥2-3/件'],
    ['N2', '蒸汽锁定按钮', '一键锁定持续出汽', '#8痛点:需持续按压', '+¥1-2/件'],
    ['N3', '硅胶握柄+隔热层', '防烫防滑', '#6痛点:握柄过热', '+¥3-5/件'],
    ['N4', '加长电源线', '≥12ft (3.6m)', '差评提及线短', '+¥2-3/件'],
    ['N5', '收纳袋', '附赠配件收纳包', '提升开箱体验', '+¥3-5/件'],
    ['N6', '双语说明书', '英文+西语', '#8痛点:说明书差', '+¥0.5-1/件'],
]

for i, row in enumerate(nice_to_have, 18):
    for j, val in enumerate(row):
        cell = ws4.cell(row=i, column=j+1, value=val)
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

apply_border(ws4, 17, 23, 5)

# Cut
ws4['A26'] = '🔴 Cut (MVP不做)'
ws4['A26'].font = Font(name='Arial', bold=True, size=12, color='C00000')

cut_headers = ['#', '功能', '砍掉原因', '', '']
for col, h in enumerate(cut_headers, 1):
    ws4.cell(row=27, column=col, value=h)
apply_header_style(ws4, 27, 5)

cut_list = [
    ['C1', 'LED照明', '23.5%覆盖率, 非核心驱动', '', ''],
    ['C2', '可调蒸汽档位', '增加成本和复杂度', '', ''],
    ['C3', '手持/拖把可转换', '对标$80+产品', '', ''],
    ['C4', '无线/电池', '技术难度高, 成本高', '', ''],
    ['C5', '大水箱≥500ml', '增加重量, 降低便携性', '', ''],
]

for i, row in enumerate(cut_list, 28):
    for j, val in enumerate(row):
        cell = ws4.cell(row=i, column=j+1, value=val)
        cell.fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')

apply_border(ws4, 27, 32, 5)

ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 22
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 28
ws4.column_dimensions['E'].width = 18

# ========================================
# Sheet 5: 测试标准
# ========================================
ws5 = wb.create_sheet('测试标准')
ws5.sheet_properties.tabColor = '7030A0'

ws5.merge_cells('A1:F1')
ws5['A1'] = 'MVP 测试标准 (Pass / Fail)'
ws5['A1'].font = title_font

headers = ['指标', '30天 Pass', '30天 Fail', '60天 Pass', '90天 Pass', '数据来源']
for col, h in enumerate(headers, 1):
    ws5.cell(row=3, column=col, value=h)
apply_header_style(ws5, 3, 6)

test_data = [
    ['日均单量', '≥10单', '<5单', '≥15单', '≥20单', 'Seller Central'],
    ['BSR排名', 'Steamers Top 20', '>Top 50', 'Top 10', 'Top 7', 'Amazon'],
    ['转化率', '≥10%', '<6%', '≥12%', '≥14%', 'Seller Central'],
    ['平均星级', '≥4.0', '<3.5', '≥4.2', '≥4.3', 'Amazon'],
    ['退货率', '≤10%', '>15%', '≤8%', '≤6%', 'Seller Central'],
    ['ACOS', '≤35%', '>50%', '≤30%', '≤25%', 'Ad Console'],
    ['TACOS', '≤25%', '>35%', '≤20%', '≤15%', '计算'],
    ['漏水差评占比', '≤5%', '>10%', '≤3%', '≤2%', 'Review分析'],
]

for i, row in enumerate(test_data, 4):
    for j, val in enumerate(row):
        cell = ws5.cell(row=i, column=j+1, value=val)
        if j == 1:
            cell.fill = green_fill
        elif j == 2:
            cell.fill = red_fill
        elif j in [3, 4]:
            cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

apply_border(ws5, 3, 11, 6)

# Kill/Continue/Pivot
ws5['A14'] = '30天检查点决策框架'
ws5['A14'].font = bold_font

kcp_headers = ['决策', '条件', '行动']
for col, h in enumerate(kcp_headers, 1):
    ws5.cell(row=15, column=col, value=h)
apply_header_style(ws5, 15, 3)

kcp_data = [
    ['✅ Continue', '日均≥10 AND 星级≥4.0 AND 退货率≤10%', '追加库存, 加大广告'],
    ['🔄 Pivot', '日均5-10 OR 星级3.5-4.0 OR 退货率10-15%', '分析原因, 优化Listing或产品'],
    ['🛑 Kill', '日均<5 AND 星级<3.5 AND 退货率>15%', '清仓处理, 止损退出'],
]

fills2 = [green_fill, yellow_fill, red_fill]
for i, (row, fill) in enumerate(zip(kcp_data, fills2), 16):
    for j, val in enumerate(row):
        cell = ws5.cell(row=i, column=j+1, value=val)
        cell.fill = fill

apply_border(ws5, 15, 18, 3)

for col in range(1, 7):
    ws5.column_dimensions[get_column_letter(col)].width = 22

# ========================================
# Sheet 6: 关键词策略
# ========================================
ws6 = wb.create_sheet('关键词策略')
ws6.sheet_properties.tabColor = '00B0F0'

ws6.merge_cells('A1:F1')
ws6['A1'] = '关键词投放策略'
ws6['A1'].font = title_font

headers = ['关键词', '月搜索量', 'CPC ($)', '优先级', '匹配类型', '备注']
for col, h in enumerate(headers, 1):
    ws6.cell(row=3, column=col, value=h)
apply_header_style(ws6, 3, 6)

kw_data = [
    ['steam cleaner', 530968, 1.53, '高竞争', 'Broad', '大词试水'],
    ['steamer cleaner', 18545, 1.40, '中等', 'Exact', '品类词'],
    ['steam cleaner for walls', 17742, 1.68, '中等', 'Exact', '场景词'],
    ['wall steamer to clean walls', 16066, 1.11, '低CPC', 'Exact', '精准场景'],
    ['vacuum steamer', 9469, 1.84, '品类相关', 'Exact', '主关键词'],
    ['steam cleaner for floors', 5932, 1.10, '中等', 'Exact', '场景词'],
    ['home steam cleaner', 4585, 3.17, 'CPC偏高', 'Phrase', '监控ROI'],
    ['tile and grout steam cleaner', 2264, 1.17, '精准', 'Exact', '高转化'],
    ['steam cleaner for home use', 1068, 1.74, '精准', 'Exact', '高转化'],
    ['steam cleaner with attachments', 219, 1.72, '精准长尾', 'Exact', '差异化匹配'],
]

for i, row in enumerate(kw_data, 4):
    for j, val in enumerate(row):
        cell = ws6.cell(row=i, column=j+1, value=val)
        if j == 1:
            cell.number_format = '#,##0'
        elif j == 2:
            cell.number_format = '$#,##0.00'

apply_border(ws6, 3, 13, 6)

ws6.column_dimensions['A'].width = 35
ws6.column_dimensions['B'].width = 14
ws6.column_dimensions['C'].width = 10
ws6.column_dimensions['D'].width = 14
ws6.column_dimensions['E'].width = 12
ws6.column_dimensions['F'].width = 18

# ========================================
# 保存
# ========================================
output_path = r'C:\Users\Administrator\vacuum_steamer_research\output\20260321_US_handheld-steam-cleaner_MVP蓝图_财务模型.xlsx'
wb.save(output_path)
print(f'Excel saved to: {output_path}')
