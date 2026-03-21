import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hfl = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
sf = Font(name='Arial', bold=True, size=10)
nf = Font(name='Arial', size=10)
bf = Font(name='Arial', bold=True, size=10)
gf = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
rf = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yf = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
sf2 = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
tf = Font(name='Arial', bold=True, size=14, color='059669')
bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

def hdr(ws, r, mc):
    for c in range(1, mc+1):
        cl = ws.cell(row=r, column=c)
        cl.font = hf; cl.fill = hfl; cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); cl.border = bd

def brd(ws, r1, r2, mc):
    for r in range(r1, r2+1):
        for c in range(1, mc+1):
            ws.cell(row=r, column=c).border = bd; ws.cell(row=r, column=c).font = nf

# ===== Sheet 1: 用户影响力排序 =====
ws1 = wb.active
ws1.title = '用户影响力排序'
ws1.sheet_properties.tabColor = '059669'
ws1.merge_cells('A1:H1')
ws1['A1'] = 'MVP 改良/创新点 — 用户影响力4维度加权排序'; ws1['A1'].font = tf

h = ['排名', '项目', '分类', 'A可感知(35%)', 'B决策力(30%)', 'C体验力(20%)', 'D成本效率(15%)', '总分']
for i, v in enumerate(h, 1): ws1.cell(row=3, column=i, value=v)
hdr(ws1, 3, 8)

data = [
    ['★1', 'LED 三色就绪灯', '进攻', 5, 4, 4, 4, 4.35],
    ['★2', '蒸汽锁定按钮', '进攻', 4, 4, 5, 4, 4.15],
    ['★3', 'Snap-Lock配件(橙色按钮)', '防守', 4, 4, 5, 3, 4.00],
    ['★4', '薄荷绿配色', '进攻', 5, 4, 2, 4, 3.95],
    ['★5', '透明水箱窗+水位刻度', '进攻', 5, 3, 4, 3, 3.90],
    ['★6', '硅胶握柄(撞色白)', '防守', 5, 3, 4, 3, 3.90],
    ['★7', '配件收纳底座', '进攻', 5, 4, 3, 3, 3.90],
    ['8', '15件配件+收纳袋', '进攻', 4, 4, 3, 3, 3.65],
    ['9', '旋锁安全盖', '防守', 3, 3, 5, 3, 3.50],
    ['10', '哑光微纹理', '进攻', 4, 3, 1, 3, 2.95],
    ['11', '加长电源线12ft', '防守', 3, 2, 3, 3, 2.75],
    ['12', '气液分离挡板', '隐形', 1, 2, 5, 3, 2.60],
    ['13', '双语说明书+QR', '防守', 2, 1, 3, 5, 2.50],
    ['14', 'EPDM密封圈+双圈', '隐形', 1, 1, 5, 4, 2.50],
    ['15', 'NTC浪涌限流器', '隐形', 1, 1, 4, 4, 2.25],
    ['16', 'PP+30%GF壳体', '隐形', 1, 1, 4, 3, 2.10],
    ['17', '可复位温控器', '隐形', 1, 1, 4, 3, 2.10],
    ['18', '温控死区缩小', '隐形', 1, 1, 3, 4, 2.00],
]
fills = {'进攻': gf, '防守': PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'), '隐形': rf}
for i, row in enumerate(data, 4):
    for j, v in enumerate(row):
        cl = ws1.cell(row=i, column=j+1, value=v)
        if j == 2 and v in fills: cl.fill = fills[v]
        if j == 7: cl.number_format = '0.00'
    if i <= 10: # MVP入选项加粗
        for j in range(8): ws1.cell(row=i, column=j+1).font = bf
brd(ws1, 3, 21, 8)
for c, w in [(1,6),(2,28),(3,8),(4,14),(5,14),(6,14),(7,14),(8,8)]: ws1.column_dimensions[get_column_letter(c)].width = w

# ===== Sheet 2: 成本分析 =====
ws2 = wb.create_sheet('成本分析'); ws2.sheet_properties.tabColor = 'F97316'
ws2.merge_cells('A1:E1'); ws2['A1'] = '单件成本明细 (精简版 MVP)'; ws2['A1'].font = tf
for i, v in enumerate(['成本项', '金额USD', '占售价%', '说明', '来源'], 1): ws2.cell(row=3, column=i, value=v)
hdr(ws2, 3, 5)
costs = [['采购(FOB)+改良', 13.75, '', '¥82+¥17改良=¥99, 汇率7.2', '1688+推测'],
    ['头程(海运)', 2.50, '', '~$150/CBM', '推测'], ['FBA配送', 5.80, '', '标准尺寸', '推测'],
    ['FBA仓储', 0.50, '', '月均', '推测'], ['佣金15%', 6.00, '', '15%×$39.99', 'Amazon'],
    ['广告分摊', 4.00, '', 'TACOS 10%', '推测'], ['其他', 1.50, '', '退货/包装/保险', '推测'],
    ['总成本', 34.05, '', '', ''], ['', '', '', '', ''],
    ['售价$39.99 毛利', 5.94, '14.9%', '', ''], ['售价$42.99 毛利', 8.94, '20.8%', '', ''], ['售价$44.99 毛利', 10.94, '24.3%', '', '']]
for i, row in enumerate(costs, 4):
    for j, v in enumerate(row): cl = ws2.cell(row=i, column=j+1, value=v); cl.border = bd
    if row[0] in ['总成本']: [setattr(ws2.cell(row=i, column=j+1), 'font', bf) for j in range(5)]; [setattr(ws2.cell(row=i, column=j+1), 'fill', sf2) for j in range(5)]
    if '毛利' in str(row[0]): [setattr(ws2.cell(row=i, column=j+1), 'fill', gf) for j in range(5)]
for c, w in [(1,22),(2,12),(3,10),(4,35),(5,15)]: ws2.column_dimensions[get_column_letter(c)].width = w

# ===== Sheet 3: 盈亏平衡 =====
ws3 = wb.create_sheet('盈亏平衡'); ws3.sheet_properties.tabColor = 'EF4444'
ws3.merge_cells('A1:G1'); ws3['A1'] = '12个月财务预测 (基准: $42.99, 日均15单)'; ws3['A1'].font = tf
for i, v in enumerate(['月份', '售价', '日均', '月销', '月销额', '月毛利', '累计毛利'], 1): ws3.cell(row=3, column=i, value=v)
hdr(ws3, 3, 7)
months = [[1,34.99,8,240,8398,-1362,-24987],[2,37.99,12,360,13676,1333,-23654],[3,39.99,16,480,19195,2851,-20803],
    [4,42.99,18,540,23215,4826,-15977],[5,42.99,20,600,25794,5362,-10615],[6,44.99,22,660,29693,7213,-3402],
    [7,44.99,24,720,32393,7885,4483],[8,44.99,25,750,33743,8222,12705],[9,44.99,26,780,35092,8558,21263],
    [10,44.99,28,840,37792,9231,30494],[11,44.99,30,900,40491,9903,40397],[12,44.99,30,900,40491,9903,50300]]
for i, row in enumerate(months, 4):
    for j, v in enumerate(row):
        cl = ws3.cell(row=i, column=j+1, value=v); cl.border = bd
        if j == 1: cl.number_format = '$#,##0.00'
        elif j >= 4: cl.number_format = '$#,##0'
    if row[6] >= 0: ws3.cell(row=i, column=7).fill = gf
    else: ws3.cell(row=i, column=7).fill = rf
for c in range(1, 8): ws3.column_dimensions[get_column_letter(c)].width = 14

# ===== Sheet 4: 预算方案 =====
ws4 = wb.create_sheet('预算方案'); ws4.sheet_properties.tabColor = '6366F1'
ws4.merge_cells('A1:D1'); ws4['A1'] = '三档预算方案对比'; ws4['A1'].font = tf
for i, v in enumerate(['', '精简版(MVP推荐)', '标准版', '完整版(V2)'], 1): ws4.cell(row=3, column=i, value=v)
hdr(ws4, 3, 4)
rows = [['改良项数', '7项', '11项', '18项'], ['增量成本', '+¥15-22', '+¥25-35', '+¥40-55'],
    ['采购单价', '¥99 ($13.75)', '¥112 ($15.56)', '¥129 ($17.92)'],
    ['$39.99毛利率', '14.9%', '9.4%', '2.1%'], ['$42.99毛利率', '20.8%', '15.4%', '8.4%'],
    ['$44.99毛利率', '24.3%', '19.2%', '12.5%'], ['首批投入', '$23,625', '$27,800', '$33,200'],
    ['回本周期', '6个月', '8个月', '12个月+'], ['推荐度', '★★★★★ MVP', '★★★ 有余力时', '★ V2再做']]
for i, row in enumerate(rows, 4):
    for j, v in enumerate(row):
        cl = ws4.cell(row=i, column=j+1, value=v); cl.border = bd
        if j == 1: cl.fill = gf
for c, w in [(1,16),(2,22),(3,22),(4,22)]: ws4.column_dimensions[get_column_letter(c)].width = w

# ===== Sheet 5: 测试标准 =====
ws5 = wb.create_sheet('测试标准'); ws5.sheet_properties.tabColor = '8B5CF6'
ws5.merge_cells('A1:F1'); ws5['A1'] = 'Pass/Fail 测试标准'; ws5['A1'].font = tf
for i, v in enumerate(['指标', '30天Pass', '30天Fail', '60天Pass', '90天Pass', '来源'], 1): ws5.cell(row=3, column=i, value=v)
hdr(ws5, 3, 6)
tests = [['日均单量','≥10','<5','≥15','≥20','Seller Central'],['BSR','Top20','>Top50','Top10','Top7','Amazon'],
    ['转化率','≥10%','<6%','≥12%','≥14%','Seller Central'],['星级','≥4.0','<3.5','≥4.2','≥4.3','Amazon'],
    ['退货率','≤10%','>15%','≤8%','≤6%','Seller Central'],['ACOS','≤35%','>50%','≤30%','≤25%','Ad Console'],
    ['喷水差评占比','≤5%','>10%','≤3%','≤2%','Review分析']]
for i, row in enumerate(tests, 4):
    for j, v in enumerate(row):
        cl = ws5.cell(row=i, column=j+1, value=v); cl.border = bd
        if j == 1: cl.fill = gf
        elif j == 2: cl.fill = rf
for c in range(1, 7): ws5.column_dimensions[get_column_letter(c)].width = 16

# ===== Sheet 6: 供应商需求表 =====
ws6 = wb.create_sheet('供应商需求表'); ws6.sheet_properties.tabColor = '06B6D4'
ws6.merge_cells('A1:D1'); ws6['A1'] = '供应商技术改良需求表 (一页纸)'; ws6['A1'].font = tf
for i, v in enumerate(['优先级', '项目', '精确要求', '不接受'], 1): ws6.cell(row=3, column=i, value=v)
hdr(ws6, 3, 4)
reqs = [['★1', 'LED就绪灯', '三色LED(蓝/绿/红), 贴片式, 呼吸闪烁', '无指示灯'],
    ['★2', '蒸汽锁定', '触发阀+滑动锁定, 推到位卡住=持续出汽', '仅弹簧复位'],
    ['★3', 'Snap-Lock', 'POM卡爪, ≥50N, 橙色侧面释放按钮', 'push-fit'],
    ['★4', '配色', 'Pantone 337C薄荷绿, 哑光VDI 27-30, 白色点缀', '灰/黑色'],
    ['★5', '透明水箱', 'PC透明窗40×15mm, MAX 300ml刻蚀水位线', '不透明'],
    ['★6', '气液分离', '蒸汽出口3层不锈钢多孔挡板, 孔径1-2mm', '无挡板直出']]
for i, row in enumerate(reqs, 4):
    for j, v in enumerate(row): cl = ws6.cell(row=i, column=j+1, value=v); cl.border = bd
for c, w in [(1,8),(2,14),(3,50),(4,18)]: ws6.column_dimensions[get_column_letter(c)].width = w

out = r'C:\Users\Administrator\vacuum_steamer_research\output\20260321_US_handheld-steam-cleaner_MVP蓝图_财务模型_v3.xlsx'
wb.save(out)
print(f'Excel saved: {out}')
