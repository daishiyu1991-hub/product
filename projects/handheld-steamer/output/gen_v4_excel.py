import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
MT = '059669'
HF = PatternFill(start_color=MT, end_color=MT, fill_type='solid')
HN = Font(bold=True, color='FFFFFF', size=11, name='Arial')
TF = Font(bold=True, size=14, color=MT, name='Arial')
NF = Font(name='Arial', size=10)
BF = Font(name='Arial', bold=True, size=10)
GF = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
BL = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
RF = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
YF = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
BD = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
CT = Alignment(horizontal='center',vertical='center',wrap_text=True)
LW = Alignment(horizontal='left',vertical='center',wrap_text=True)

def hdr(ws,r,mc):
    for c in range(1,mc+1):
        cl=ws.cell(row=r,column=c);cl.font=HN;cl.fill=HF;cl.alignment=CT;cl.border=BD

def dc(cl,b=False):
    cl.border=BD;cl.font=BF if b else NF;cl.alignment=CT

# ===== Sheet 1: 用户画像 =====
ws1=wb.active;ws1.title='用户画像';ws1.sheet_properties.tabColor=MT
ws1.merge_cells('A1:B1');ws1['A1']='Primary Persona — Sarah（28-42岁妈妈）';ws1['A1'].font=TF
h=['维度','详细描述']
for i,v in enumerate(h,1):ws1.cell(row=3,column=i,value=v)
hdr(ws1,3,2)
data=[
    ['人物标签','28-42岁女性，已婚有1-2个孩子(至少1个≤5岁)和/或宠物'],
    ['核心诉求','日常轻量清洁，不想用化学清洁剂，有小孩宠物安全顾虑'],
    ['购买决策','TikTok看到清洁视频 → Amazon搜索 → 按价格筛$20-50 → 看主图+价格 → 3分钟内下单（冲动型）'],
    ['审美偏好','清新家居感，马卡龙色系，"好看到不用藏起来"的小家电'],
    ['价格敏感度','$30-45舒适区, >$50犹豫, 会和CURMEDI $43.99对比'],
    ['痛点优先级','① 安全(漏水烫到小孩) > ② 好用(不喷水/配件好装) > ③ 效果(日常够用就行)'],
    ['使用场景','厨房灶台油渍 > 浴室瓷砖缝 > 偶尔沙发/窗户'],
    ['信息触达','TikTok清洁hack → Amazon搜索 → Vine/视频评价'],
    ['绝对不接受','漏水烫手 / 配件弹飞 / 用3次就坏'],
]
for i,row in enumerate(data,4):
    for j,v in enumerate(row):cl=ws1.cell(row=i,column=j+1,value=v);dc(cl)
r=4+len(data)+1
ws1.cell(row=r,column=1,value='Secondary Persona — Alex（22-35岁租房青年）');ws1.cell(row=r,column=1).font=Font(bold=True,size=12,color=MT)
data2=[['标签','22-35岁,单身,公寓租户'],['诉求','退租大扫除/春季深度清洁'],['决策','Google比较文章→Amazon→仔细看差评（研究型）'],['场景','浴室瓷砖缝>厨房>窗户>车内']]
for i,v in enumerate(['维度','描述'],1):ws1.cell(row=r+1,column=i,value=v)
hdr(ws1,r+1,2)
for i,row in enumerate(data2,r+2):
    for j,v in enumerate(row):cl=ws1.cell(row=i,column=j+1,value=v);dc(cl)
# 决策影响
dr=r+2+len(data2)+1
ws1.cell(row=dr,column=1,value='画像对产品决策的影响');ws1.cell(row=dr,column=1).font=Font(bold=True,size=12,color=MT)
for i,v in enumerate(['决策点','Sarah指导','最终决策'],1):ws1.cell(row=dr+1,column=i,value=v)
hdr(ws1,dr+1,3)
dd=[['配色','清新马卡龙→薄荷绿','薄荷绿 Pantone 337C'],['LED灯','冲动购买→主图吸引','✅ 做'],['蒸汽弱','日常够用→管理期望','Listing写daily cleaning'],['第一卖点','安全第一','SAFE for your family'],['定价起步','$34.99低于CURMEDI','$34.99+20%Coupon']]
for i,row in enumerate(dd,dr+2):
    for j,v in enumerate(row):cl=ws1.cell(row=i,column=j+1,value=v);dc(cl)
ws1.column_dimensions['A'].width=18;ws1.column_dimensions['B'].width=75;ws1.column_dimensions['C'].width=30

# ===== Sheet 2: 用户影响力排序 =====
ws2=wb.create_sheet('影响力排序');ws2.sheet_properties.tabColor=MT
ws2.merge_cells('A1:I1');ws2['A1']='改良/创新项 — 用户影响力4维度加权排序（Sarah权重：可感知40%）';ws2['A1'].font=TF
h2=['排名','项目','分类','A可感知(40%)','B决策力(25%)','C体验力(20%)','D成本效率(15%)','总分','成本']
for i,v in enumerate(h2,1):ws2.cell(row=3,column=i,value=v)
hdr(ws2,3,9)
items=[
    ['★1','LED 三色就绪灯','进攻',5,4,4,4,4.40,'+¥2-3'],
    ['★2','薄荷绿配色','进攻',5,4,2,4,4.00,'+¥1-2'],
    ['★3','Snap-Lock配件(橙色按钮)','防守',4,4,5,3,4.05,'+¥3-5'],
    ['★4','蒸汽锁定按钮','进攻',4,3,5,4,3.95,'+¥1-2'],
    ['★5','透明水箱窗+水位刻度','进攻',5,3,4,3,3.95,'+¥2-3'],
    ['★6','硅胶握柄(白色撞色)','防守',5,3,4,3,3.95,'+¥3-5'],
    ['★7','气液分离挡板(Anti-Spit)','隐形',1,2,5,3,2.50,'+¥3-5'],
    ['—','以下砍掉→V2','','','','','','',''],
    ['8','15件配件+收纳袋','进攻',4,4,3,3,3.65,'+¥3-5'],
    ['9','旋锁安全盖','防守',3,3,5,3,3.45,'+¥3-5'],
    ['10','加长电源线12ft','防守',3,2,3,3,2.75,'+¥2-3'],
    ['14','EPDM密封圈','隐形',1,1,5,4,2.40,'+¥1.5-3'],
    ['16','PP+30%GF壳体','隐形',1,1,4,3,2.05,'+¥2-4'],
]
fills={'进攻':GF,'防守':BL,'隐形':RF}
for i,row in enumerate(items,4):
    for j,v in enumerate(row):
        cl=ws2.cell(row=i,column=j+1,value=v);dc(cl,b=(i<=10))
    cat=row[2]
    if cat in fills:
        for j in range(9):ws2.cell(row=i,column=j+1).fill=fills[cat]
    if row[0]=='—':
        for j in range(9):ws2.cell(row=i,column=j+1).font=Font(color='999999',size=10)
for c,w in [(1,6),(2,32),(3,8),(4,13),(5,13),(6,13),(7,13),(8,8),(9,10)]:ws2.column_dimensions[get_column_letter(c)].width=w

# ===== Sheet 3: 成本分析 =====
ws3=wb.create_sheet('成本分析');ws3.sheet_properties.tabColor=MT
ws3.merge_cells('A1:D1');ws3['A1']='单件成本明细（精简版 7项改良）';ws3['A1'].font=TF
for i,v in enumerate(['成本项','金额USD','占售价%','来源'],1):ws3.cell(row=3,column=i,value=v)
hdr(ws3,3,4)
costs=[['采购(¥82+¥19改良=¥101)',14.03,'','1688+推测'],['头程',2.50,'','推测'],['FBA配送',5.80,'','推测'],['FBA仓储',0.50,'','推测'],['佣金15%',6.00,'','Amazon'],['广告TACOS10%',4.00,'','推测'],['其他',1.50,'','推测'],['总成本',34.33,'',''],['','','',''],['$39.99毛利',5.66,'14.2%',''],['$42.99毛利',8.66,'20.1%',''],['$44.99毛利',10.66,'23.7%','']]
for i,row in enumerate(costs,4):
    for j,v in enumerate(row):cl=ws3.cell(row=i,column=j+1,value=v);dc(cl)
    if '总成本' in str(row[0]):
        for j in range(4):ws3.cell(row=i,column=j+1).font=BF
    if '毛利' in str(row[0]):
        for j in range(4):ws3.cell(row=i,column=j+1).fill=GF
for c,w in [(1,28),(2,12),(3,10),(4,15)]:ws3.column_dimensions[get_column_letter(c)].width=w

# ===== Sheet 4: 盈亏平衡 =====
ws4=wb.create_sheet('盈亏平衡');ws4.sheet_properties.tabColor=MT
ws4.merge_cells('A1:H1');ws4['A1']='12个月财务预测（基准$42.99, 日均15单）';ws4['A1'].font=TF
for i,v in enumerate(['月份','售价','日均','月销','月销额','月毛利','累计毛利','备注'],1):ws4.cell(row=3,column=i,value=v)
hdr(ws4,3,8)
ms=[[1,34.99,8,240,8398,-1362,-1362,'冲量期'],[2,37.99,12,360,13676,1333,-29,'爬坡'],[3,39.99,15,450,17996,2851,2822,'转化提升'],
    [4,42.99,17,510,21925,4826,7648,'提价'],[5,42.99,18,540,23215,5362,13010,'稳定'],[6,44.99,20,600,26994,7213,20223,'加速'],
    [7,44.99,22,660,29693,7885,28108,'稳定'],[8,44.99,24,720,32393,8558,36666,'备PrimeDay'],[9,44.99,25,750,33743,8894,45560,'旺季'],
    [10,44.99,26,780,35092,9231,54791,'增长'],[11,44.99,28,840,37792,9903,64694,'黑五'],[12,44.99,30,900,40491,10574,75268,'圣诞']]
for i,row in enumerate(ms,4):
    for j,v in enumerate(row):
        cl=ws4.cell(row=i,column=j+1,value=v);dc(cl)
        if j==1 and isinstance(v,float):cl.number_format='$#,##0.00'
        if j in [4,5,6] and isinstance(v,(int,float)):cl.number_format='$#,##0'
    if row[6]>=0:ws4.cell(row=i,column=7).fill=GF
    else:ws4.cell(row=i,column=7).fill=RF
for c in range(1,9):ws4.column_dimensions[get_column_letter(c)].width=14

# ===== Sheet 5: 测试标准 =====
ws5=wb.create_sheet('测试标准');ws5.sheet_properties.tabColor=MT
ws5.merge_cells('A1:F1');ws5['A1']='Pass/Fail 测试标准（含安全类差评）';ws5['A1'].font=TF
for i,v in enumerate(['指标','30天Pass','30天Fail','60天Pass','90天Pass','来源'],1):ws5.cell(row=3,column=i,value=v)
hdr(ws5,3,6)
ts=[['日均单量','≥10','<5','≥15','≥20','Seller Central'],['转化率','≥10%','<6%','≥12%','≥14%','Seller Central'],
    ['星级','≥4.0','<3.5','≥4.2','≥4.3','Amazon'],['退货率','≤10%','>15%','≤8%','≤6%','Seller Central'],
    ['ACOS','≤35%','>50%','≤30%','≤25%','Ad Console'],['BSR','Top20','>Top50','Top10','Top7','Amazon'],
    ['喷水差评占比','≤5%','>10%','≤3%','≤2%','Review'],['★安全类差评占比','≤3%','>8%','≤2%','≤1%','Review']]
for i,row in enumerate(ts,4):
    for j,v in enumerate(row):
        cl=ws5.cell(row=i,column=j+1,value=v);dc(cl)
        if j==1:cl.fill=GF
        elif j==2:cl.fill=RF
    if '安全' in row[0]:
        for j in range(6):ws5.cell(row=i,column=j+1).fill=YF;ws5.cell(row=i,column=j+1).font=BF
for c in range(1,7):ws5.column_dimensions[get_column_letter(c)].width=18

# ===== Sheet 6: 供应商需求表 =====
ws6=wb.create_sheet('供应商需求表');ws6.sheet_properties.tabColor=MT
ws6.merge_cells('A1:D1');ws6['A1']='供应商技术改良需求表（按Sarah优先级）';ws6['A1'].font=TF
for i,v in enumerate(['优先级','项目','精确要求','不接受'],1):ws6.cell(row=3,column=i,value=v)
hdr(ws6,3,4)
reqs=[['★1','Snap-Lock配件','POM卡爪, ≥50N, 橙色侧面释放按钮','push-fit'],
    ['★2','气液分离','蒸汽出口3层不锈钢多孔挡板1-2mm','无挡板直出'],
    ['★3','LED就绪灯','三色LED蓝/绿/红, 贴片式, 呼吸模式','无指示灯'],
    ['★4','薄荷绿配色','Pantone 337C, 哑光VDI 27-30, 白色点缀','灰/黑色'],
    ['★5','硅胶握柄','白色硅胶2-3mm+菱形防滑纹','裸硬塑料'],
    ['★6','透明水箱','PC透明窗40×15mm, MAX 300ml刻蚀','不透明'],
    ['★7','蒸汽锁定','触发阀滑动锁定, 推到位卡住','仅弹簧复位']]
for i,row in enumerate(reqs,4):
    for j,v in enumerate(row):cl=ws6.cell(row=i,column=j+1,value=v);dc(cl)
for c,w in [(1,8),(2,16),(3,50),(4,18)]:ws6.column_dimensions[get_column_letter(c)].width=w

out=r'C:\Users\Administrator\vacuum_steamer_research\output\20260321_US_handheld-steam-cleaner_MVP蓝图_财务模型_v4.xlsx'
wb.save(out)
print(f'Saved: {out}')
