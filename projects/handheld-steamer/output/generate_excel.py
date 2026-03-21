#!/usr/bin/env python3
"""
Vacuum Steamer 美国站市场调研数据 Excel 生成脚本
数据来源：Sorftime MCP
生成日期：2026-03-20
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()

# ============================================================
# Style definitions
# ============================================================
header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
title_font = Font(name='Microsoft YaHei', bold=True, size=14, color='1E40AF')
subtitle_font = Font(name='Microsoft YaHei', bold=True, size=10, color='64748B')
data_font = Font(name='Microsoft YaHei', size=10)
number_font = Font(name='Microsoft YaHei', size=10)
highlight_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
green_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
yellow_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')


def setup_header(ws, row, headers, col_widths=None):
    """Apply header styling to a row."""
    for col_idx, (header, width) in enumerate(zip(headers, col_widths or [15]*len(headers)), 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_data_row(ws, row, data, num_cols=None):
    """Write a data row with styling."""
    num_cols = num_cols or set()
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = data_font
        cell.border = thin_border
        if col_idx in num_cols:
            cell.alignment = right_align
        else:
            cell.alignment = left_align


def add_title(ws, row, title, subtitle=None):
    """Add a title and optional subtitle."""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = title_font
    if subtitle:
        cell = ws.cell(row=row + 1, column=1, value=subtitle)
        cell.font = subtitle_font


# ============================================================
# Sheet 1: 关键词数据
# ============================================================
ws1 = wb.active
ws1.title = '关键词数据'

add_title(ws1, 1, 'Vacuum Steamer 关键词分析数据')
ws1.cell(row=2, column=1, value='数据来源：Sorftime MCP | 报告日期：2026-03-20').font = subtitle_font

# Core keyword info
ws1.cell(row=4, column=1, value='一、核心关键词信息').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers = ['关键词', '周搜索量', '月搜索量', 'CPC推荐出价', '搜索排名', '旺季', '搜索结果竞品数', '首页非BS Top100占比', '首页低评价占比']
widths = [20, 12, 12, 14, 12, 10, 16, 18, 16]
setup_header(ws1, 5, headers, widths)

core_data = ['vacuum steamer', 1993, 9469, '$1.84', 218905, '3月', 34734, '16.67%', '100%']
write_data_row(ws1, 6, core_data, {2, 3, 5, 7})
for col in range(1, 10):
    ws1.cell(row=6, column=col).fill = highlight_fill

# Extended keywords
ws1.cell(row=8, column=1, value='二、核心延伸关键词 TOP10').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers2 = ['关键词', '周搜索量', '月搜索量', 'CPC推荐出价', '旺季', '竞争评估']
widths2 = [28, 12, 12, 14, 14, 12]
setup_header(ws1, 9, headers2, widths2)

keywords_data = [
    ['steam cleaner', 123899, 530968, '$1.53', '2月', '高竞争'],
    ['steam mop', 58133, 276501, '$1.41', '7月', '高竞争'],
    ['shark steam and scrub mop', 25711, 88584, '$0.75', '7,9,11月', '品牌词'],
    ['floor steamer', 13748, 60458, '$1.51', '均衡', '中竞争'],
    ['bissell steam cleaner', 14397, 59776, '$0.94', '3月', '品牌词'],
    ['dupray steam cleaner', 12453, 54361, '$9.58', '11月', '品牌词'],
    ['steam vacuum cleaner', 4960, 21407, '$1.99', '均衡', '中竞争'],
    ['steamer cleaner', 4415, 18545, '$1.40', '11月', '中低竞争'],
    ['vacuum steamer', 1993, 9469, '$1.84', '3月', '低竞争'],
    ['steam vacuum', 1966, 8574, '$1.23', '7,11月', '低竞争'],
]

for i, row_data in enumerate(keywords_data, 10):
    write_data_row(ws1, i, row_data, {2, 3})
    # Highlight target keywords
    if row_data[0] in ('vacuum steamer', 'steam vacuum cleaner', 'steam vacuum', 'steamer cleaner'):
        for col in range(1, 7):
            ws1.cell(row=i, column=col).fill = green_fill

# Search volume trend
ws1.cell(row=21, column=1, value='三、"vacuum steamer" 月搜索量趋势').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers3 = ['月份', '月搜索量', '环比变化', '备注']
widths3 = [12, 12, 12, 30]
setup_header(ws1, 22, headers3, widths3)

trend_data = [
    ['2024.03', 585, '-', '基准'],
    ['2024.07', 750, '+28% ⚠️推测', '⚠️ 推测值'],
    ['2024.11', 1081, '+44%', ''],
    ['2025.01', 2000, '+85% ⚠️推测', '⚠️ 推测值'],
    ['2025.02', 5687, '+184%', '爆发起点'],
    ['2025.03', 24981, '+339%', '🔥 爆发峰值'],
    ['2025.07', 19173, '-23%', '回落企稳'],
    ['2025.08', 23155, '+21%', ''],
    ['2025.11', 22660, '-2%', '高位稳定'],
    ['2026.01', 20089, '-11%', ''],
    ['2026.02', 17380, '-13%', ''],
]

for i, row_data in enumerate(trend_data, 23):
    write_data_row(ws1, i, row_data, {2})

# Brand distribution
ws1.cell(row=35, column=1, value='四、搜索首页品牌与卖家分布').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers4 = ['品牌/卖家', '占比', '类型']
widths4 = [20, 12, 12]
setup_header(ws1, 36, headers4, widths4)

brand_data = [
    ['HiLIFE', '90.31%', '品牌'],
    ['Bissell', '31.80%', '品牌'],
    ['Shark', '11.08%', '品牌'],
    ['HiLIFE Direct', '51.04%', '卖家'],
    ['Amazon', '48.62%', '卖家'],
]

for i, row_data in enumerate(brand_data, 37):
    write_data_row(ws1, i, row_data)


# ============================================================
# Sheet 2: Top产品列表
# ============================================================
ws2 = wb.create_sheet('Top产品列表')

add_title(ws2, 1, 'Vacuum Steamer 相关产品数据')
ws2.cell(row=2, column=1, value='数据来源：Sorftime MCP | 报告日期：2026-03-20').font = subtitle_font

# $20-50 products
ws2.cell(row=4, column=1, value='一、$20-50 价位段产品（用户目标区间）').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_p = ['ASIN', '品牌', '价格', '月销量', '星级', '评论数', '上架日期', '产品类型', '月销额']
widths_p = [16, 14, 10, 12, 8, 12, 14, 16, 14]
setup_header(ws2, 5, headers_p, widths_p)

products_data = [
    ['B0GFNBBBJX', 'HiLIFE', '$29.99', 57600, 4.6, 130054, '2026-01', '衣物挂烫机', '$1,727,424'],
    ['B0DCF593FL', 'HiLIFE', '$42.99', 57600, 4.1, 131126, '2024-08', '衣物挂烫机', '$2,476,224'],
    ['B0D1CCGJ5W', 'Tineco', '$29.99', 18635, 4.8, 15277, '2024-04', '清洁液(配件)', '$558,915'],
    ['B0CLRNZT78', 'OGHom', '$26.99', 15705, 4.3, 74, '2023-12', '衣物挂烫机', '$423,888'],
    ['B08P456NHG', 'BEAUTURAL', '$21.99', 14359, 4.3, 15903, '2020-11', '衣物挂烫机', '$315,735'],
    ['B0DGQ7L856', 'CURMEDI', '$43.99', 13431, 4.1, 1003, '2024-10', '手持蒸汽清洁器', '$590,830'],
]

for i, row_data in enumerate(products_data, 6):
    write_data_row(ws2, i, row_data, {4, 6, 9})
    if row_data[7] == '手持蒸汽清洁器':
        for col in range(1, 10):
            ws2.cell(row=i, column=col).fill = green_fill

# CURMEDI details
ws2.cell(row=13, column=1, value='二、重点竞品详情 - CURMEDI B0DGQ7L856').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

detail_items = [
    ('ASIN', 'B0DGQ7L856'),
    ('品牌', 'CURMEDI'),
    ('价格', '$43.99'),
    ('月销量', '13,431'),
    ('月销额', '$590,830'),
    ('星级', '4.1'),
    ('评价数', '1,003'),
    ('上架日期', '2024-10-10'),
    ('卖家', 'YUTING US'),
    ('来源', 'CN（中国）'),
    ('所属类目', 'Handheld Steamers'),
    ('类目排名', '#7'),
    ('产品特点', '12件配件套装，1050W功率，3-5分钟加热'),
]

for i, (key, value) in enumerate(detail_items, 14):
    ws2.cell(row=i, column=1, value=key).font = Font(name='Microsoft YaHei', bold=True, size=10)
    ws2.cell(row=i, column=1).border = thin_border
    ws2.cell(row=i, column=2, value=value).font = data_font
    ws2.cell(row=i, column=2).border = thin_border

# Steam Mops Top5
ws2.cell(row=28, column=1, value='三、Steam Mops 类目 Top5 头部产品').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_top = ['排名', 'ASIN', '品牌', '月销量', '价格', '星级', '评价数', '上架日期']
widths_top = [8, 16, 14, 12, 10, 8, 12, 14]
setup_header(ws2, 29, headers_top, widths_top)

top5_data = [
    [1, 'B0D1VZ9SQ4', 'Shark', 16864, '$159', 4.4, 1918, '2024-05'],
    [2, 'B0768ZC23W', 'Shark', 16713, 'N/A', 4.4, 18954, '2017-10'],
    [3, 'B07XZYLMK1', 'McCULLOCH', 8108, '$139.99', 4.3, 43951, '2019-10'],
    [4, 'B0091YYUAM', 'Bissell', 7296, 'N/A', 4.5, 48087, '2012-10'],
    [5, 'B0CJ4F8WYK', 'Bissell', 7017, '$119.99', 4.0, 1923, '2024-01'],
]

for i, row_data in enumerate(top5_data, 30):
    write_data_row(ws2, i, row_data, {1, 4, 7})

# Product features
ws2.cell(row=36, column=1, value='四、产品特征维度覆盖率').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_feat = ['排名', '特征维度', '产品覆盖率', '重要性', '建议']
widths_feat = [8, 24, 14, 12, 16]
setup_header(ws2, 37, headers_feat, widths_feat)

features_data = [
    [1, '无化学品清洁 (Chemical-Free)', '100%', '⭐⭐⭐⭐⭐', '必备'],
    [2, '多表面适用 (Multi-Surface)', '94.1%', '⭐⭐⭐⭐⭐', '必备'],
    [3, '快速加热 10秒内 (Quick Heat-Up)', '64.7%', '⭐⭐⭐⭐', '高优'],
    [4, '大容量水箱 350ml+ (Large Tank)', '58.8%', '⭐⭐⭐⭐', '高优'],
    [5, '可拆卸/可更换布垫 (Removable Pads)', '52.9%', '⭐⭐⭐', '中优'],
    [6, '轻便/便携 (Lightweight)', '47.1%', '⭐⭐⭐', '中优'],
    [7, '多功能配件10件+ (Accessories)', '47.1%', '⭐⭐⭐', '中优'],
    [8, '手持式设计 (Handheld)', '41.2%', '⭐⭐⭐', '差异化'],
    [9, '可调蒸汽强度 (Adjustable Steam)', '35.3%', '⭐⭐⭐', '差异化'],
    [10, '长电源线 18"+ (Long Cord)', '35.3%', '⭐⭐', '加分项'],
]

for i, row_data in enumerate(features_data, 38):
    write_data_row(ws2, i, row_data, {1})


# ============================================================
# Sheet 3: 类目趋势数据
# ============================================================
ws3 = wb.create_sheet('类目趋势数据')

add_title(ws3, 1, '类目趋势分析数据')
ws3.cell(row=2, column=1, value='数据来源：Sorftime MCP | 报告日期：2026-03-20').font = subtitle_font

# Category overview
ws3.cell(row=4, column=1, value='一、相关类目概览对比').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_cat = ['指标', 'Steam Mops (3303861011)', 'Stick Vacuums (510112)']
widths_cat = [22, 24, 24]
setup_header(ws3, 5, headers_cat, widths_cat)

cat_compare = [
    ['Top100 月销量', '116,727', '306,137'],
    ['Top100 月销额', '$10,963,659', '$41,528,238'],
    ['均价', '$83.77', '$140.10'],
    ['均星级', '4.27', '4.35'],
    ['均评价数', '2,575', '4,840'],
    ['Top3 产品销量占比', '35.71%', '17.65%'],
    ['Top3 品牌占比', '54.49%', '45.53%'],
    ['新品(3个月内)销量占比', '6.50%', '6.59%'],
    ['中国卖家占比', '76.36%', '85.71%'],
    ['亚马逊自营占比', '61.56%', '50.61%'],
    ['旺季', '均衡', '7月'],
]

for i, row_data in enumerate(cat_compare, 6):
    write_data_row(ws3, i, row_data)

# Sales trend
ws3.cell(row=18, column=1, value='二、Steam Mops 类目月销量趋势').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_trend = ['月份', '月销量', '同比变化', '备注']
widths_trend = [12, 14, 14, 20]
setup_header(ws3, 19, headers_trend, widths_trend)

sales_trend = [
    ['2024.03', 84635, '-', '基准年'],
    ['2024.07', 112257, '-', ''],
    ['2024.11', 116562, '-', ''],
    ['2025.03', 110602, '+30.7% YoY', ''],
    ['2025.07', 160871, '+43.3% YoY', '🔥 峰值'],
    ['2025.11', 127661, '+9.5% YoY', ''],
    ['2026.01', 96274, '-', ''],
    ['2026.02', 104041, '+8.1% MoM', ''],
    ['2026.03', 61880, '-', '月中数据'],
]

for i, row_data in enumerate(sales_trend, 20):
    write_data_row(ws3, i, row_data, {2})

# New product share trend
ws3.cell(row=30, column=1, value='三、新品销量占比趋势 (Steam Mops)').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_new = ['月份', '新品销量占比', '趋势', '分析']
widths_new = [12, 16, 8, 30]
setup_header(ws3, 31, headers_new, widths_new)

new_trend = [
    ['2024.03', '11.82%', '—', '基准水平'],
    ['2024.06', '8.38%', '↓', '低谷'],
    ['2025.03', '11.90%', '→', '恢复基准'],
    ['2025.06', '21.03%', '↑↑', '突破20%'],
    ['2025.08', '27.52%', '↑↑', '高位'],
    ['2026.01', '31.04%', '↑', '峰值 - 新品机会最大'],
    ['2026.03', '26.73%', '→', '高位稳定'],
]

for i, row_data in enumerate(new_trend, 32):
    write_data_row(ws3, i, row_data)
    if float(row_data[1].rstrip('%')) > 25:
        for col in range(1, 5):
            ws3.cell(row=i, column=col).fill = green_fill

# Avg price trend
ws3.cell(row=40, column=1, value='四、均价趋势 (Steam Mops)').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_price = ['月份', '均价($)', '趋势']
widths_price = [12, 12, 8]
setup_header(ws3, 41, headers_price, widths_price)

price_trend = [
    ['2024.03', 93.10, '—'],
    ['2024.07', 87.04, '↓'],
    ['2025.01', 100.63, '↑ 峰值'],
    ['2025.08', 103.31, '↑'],
    ['2025.11', 91.34, '↓'],
    ['2026.03', 83.03, '↓ 当前'],
]

for i, row_data in enumerate(price_trend, 42):
    write_data_row(ws3, i, row_data, {2})


# ============================================================
# Sheet 4: 竞品差评分析
# ============================================================
ws4 = wb.create_sheet('竞品差评分析')

add_title(ws4, 1, '竞品差评痛点分析')
ws4.cell(row=2, column=1, value='数据来源：Sorftime MCP 差评分析 | 报告日期：2026-03-20').font = subtitle_font

# CURMEDI
ws4.cell(row=4, column=1, value='一、CURMEDI B0DGQ7L856 差评痛点（手持蒸汽清洁器 $43.99）').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_pain = ['痛点', '典型差评原文', '严重程度', '改进方向', '产品设计建议']
widths_pain = [16, 36, 12, 20, 26]
setup_header(ws4, 5, headers_pain, widths_pain)

curmedi_pain = [
    ['电路跳闸/功率过大', 'cuts the power in my entire apartment', '🔴 严重', '降低功率至900W以下', '加电流缓启动电路'],
    ['漏水/吐水', 'Spits water doesn\'t work right', '🔴 严重', '优化密封圈设计', '增加防滴漏阀门'],
    ['配件松动/不牢固', 'attachments don\'t fit properly', '🟡 中等', '改用卡扣式连接', '增加公差控制精度'],
    ['水箱太小', 'Tank too small', '🟡 中等', '水箱≥380ml', '设计快充水口'],
    ['清洁力弱', 'barely clean anything', '🟡 中等', '提升蒸汽压力和温度', '优化蒸汽发生器'],
    ['耐用性差', 'stopped working after 5 uses', '🔴 严重', '提升核心组件品质', '延长保修至2年'],
    ['安全开关不便', 'no button to hold steam on', '🟢 轻微', '增加锁定持续出汽按钮', '人体工程学按钮设计'],
    ['蒸汽控制差', 'Too much steam escaping with no control', '🟡 中等', '增加2-3档蒸汽调节', '旋钮式蒸汽控制'],
]

for i, row_data in enumerate(curmedi_pain, 6):
    write_data_row(ws4, i, row_data)
    severity = row_data[2]
    if '严重' in severity:
        ws4.cell(row=i, column=3).fill = red_fill
    elif '中等' in severity:
        ws4.cell(row=i, column=3).fill = yellow_fill
    elif '轻微' in severity:
        ws4.cell(row=i, column=3).fill = green_fill

# HiLIFE
ws4.cell(row=15, column=1, value='二、HiLIFE B0CSYR4HRD 差评痛点（衣物挂烫机）').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_pain2 = ['痛点', '典型差评原文', '严重程度', '改进方向']
widths_pain2 = [16, 36, 12, 26]
setup_header(ws4, 16, headers_pain2, widths_pain2)

hilife_pain = [
    ['去皱效果差', 'Did not adequately remove wrinkles', '🔴 严重', '提升蒸汽量和温度'],
    ['漏水/滴水', 'Gets Wrinkles Out, but Drips Water', '🔴 严重', '防滴漏设计优化'],
    ['水箱接口脆弱', 'plastic tabs broke immediately', '🟡 中等', '改用加强材质'],
    ['耐用性差', 'only steams for about...after 6 months', '🔴 严重', '品质升级'],
    ['使用不便', 'More time consuming than an ordinary iron', '🟡 中等', '优化加热速度'],
]

for i, row_data in enumerate(hilife_pain, 17):
    write_data_row(ws4, i, row_data)
    severity = row_data[2]
    if '严重' in severity:
        ws4.cell(row=i, column=3).fill = red_fill
    elif '中等' in severity:
        ws4.cell(row=i, column=3).fill = yellow_fill

# Common pain points
ws4.cell(row=23, column=1, value='三、共性差评痛点归类').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_common = ['排名', '共性痛点', '影响产品类型', '优先级', '建议改进方案']
widths_common = [8, 18, 18, 10, 30]
setup_header(ws4, 24, headers_common, widths_common)

common_pain = [
    [1, '漏水/滴水', '蒸汽清洁器+衣物挂烫机', '🔴 P0', '优化密封圈+防滴漏阀+严格QC测试'],
    [2, '耐用性差', '蒸汽清洁器+衣物挂烫机', '🔴 P0', '核心部件升级+延长保修+老化测试'],
    [3, '功能不达预期', '蒸汽清洁器+衣物挂烫机', '🟡 P1', '提升蒸汽压力/温度+配件优化'],
]

for i, row_data in enumerate(common_pain, 25):
    write_data_row(ws4, i, row_data, {1})


# ============================================================
# Sheet 5: 1688采购价格
# ============================================================
ws5 = wb.create_sheet('1688采购价格')

add_title(ws5, 1, '1688 采购价格参考数据')
ws5.cell(row=2, column=1, value='数据来源：Sorftime MCP | 报告日期：2026-03-20 | 汇率：1 USD ≈ 7.2 CNY（⚠️推测）').font = subtitle_font

# Product types
ws5.cell(row=4, column=1, value='一、手持蒸汽清洁器 1688 价格区间').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_1688 = ['产品类型', '人民币价格区间', '美元折算', '适用目标价位', '推荐度', '备注']
widths_1688 = [22, 16, 14, 16, 10, 30]
setup_header(ws5, 5, headers_1688, widths_1688)

price_1688 = [
    ['高压蒸汽清洁器', '¥40-130', '$5.5-18', '$30-45', '✅ 推荐', '适合$35-45售价产品，毛利空间大'],
    ['多功能手持蒸汽机', '¥67-185', '$9-25', '$35-50', '✅ 推荐', '配件丰富，差异化空间'],
    ['大型2L蒸汽清洁器', '¥327-428', '$45-59', '$80+', '⚠️ 慎重', '采购成本偏高，超出目标价位'],
    ['布艺清洗一体机', '¥230-694', '$32-96', '$100+', '❌ 不推荐', '成本过高，不适合$20-50价位'],
]

for i, row_data in enumerate(price_1688, 6):
    write_data_row(ws5, i, row_data)
    if '推荐' in row_data[4] and '✅' in row_data[4]:
        for col in range(1, 7):
            ws5.cell(row=i, column=col).fill = green_fill
    elif '慎重' in row_data[4]:
        for col in range(1, 7):
            ws5.cell(row=i, column=col).fill = yellow_fill
    elif '不推荐' in row_data[4]:
        for col in range(1, 7):
            ws5.cell(row=i, column=col).fill = red_fill

# Cost estimate
ws5.cell(row=11, column=1, value='二、成本结构估算（售价 $42.99）⚠️ 推测').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_cost = ['成本项目', '金额', '占售价比例', '说明']
widths_cost = [20, 12, 14, 30]
setup_header(ws5, 12, headers_cost, widths_cost)

cost_data = [
    ['1688采购成本', '$9.00', '20.9%', '⚠️推测，基于¥65中间价'],
    ['国际运费(海运)', '$3.00', '7.0%', '⚠️推测'],
    ['FBA头程', '$2.50', '5.8%', '⚠️推测'],
    ['FBA配送费', '$6.50', '15.1%', '⚠️推测，基于产品重量和尺寸'],
    ['亚马逊佣金(15%)', '$6.45', '15.0%', '标准佣金率'],
    ['包装/标签', '$1.50', '3.5%', '⚠️推测'],
    ['总成本', '$28.95', '67.3%', ''],
    ['毛利', '$14.04', '32.7%', ''],
]

for i, row_data in enumerate(cost_data, 13):
    write_data_row(ws5, i, row_data)
    if row_data[0] == '毛利':
        for col in range(1, 5):
            ws5.cell(row=i, column=col).fill = green_fill
            ws5.cell(row=i, column=col).font = Font(name='Microsoft YaHei', bold=True, size=10)

# P&L scenarios
ws5.cell(row=22, column=1, value='三、盈亏场景测算 ⚠️ 推测').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_pnl = ['场景', '月销量', '月营收', '毛利', '广告费', '净利', '净利率']
widths_pnl = [10, 10, 14, 12, 12, 12, 10]
setup_header(ws5, 23, headers_pnl, widths_pnl)

pnl_data = [
    ['保守', 1000, '$42,990', '$14,040', '$3,000', '$11,040', '25.7%'],
    ['基准', 3000, '$128,970', '$42,120', '$5,000', '$37,120', '28.8%'],
    ['乐观', 8000, '$343,920', '$112,320', '$10,000', '$102,320', '29.7%'],
]

for i, row_data in enumerate(pnl_data, 24):
    write_data_row(ws5, i, row_data, {2})
    if row_data[0] == '基准':
        for col in range(1, 8):
            ws5.cell(row=i, column=col).fill = highlight_fill

# First batch investment
ws5.cell(row=28, column=1, value='四、首批投入资金估算 ⚠️ 推测').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_invest = ['投入项目', '金额', '备注']
widths_invest = [22, 16, 30]
setup_header(ws5, 29, headers_invest, widths_invest)

invest_data = [
    ['首批备货1000件', '$12,000', '基于1688采购价¥65/件'],
    ['国际运费+头程', '$5,500', '海运+FBA头程'],
    ['UL/ETL认证', '$3,000-8,000', '电器安全认证（必要）'],
    ['品牌注册+商标', '$500', ''],
    ['Listing制作(A+)', '$500', '包含摄影和A+内容'],
    ['广告启动资金(3个月)', '$9,000-15,000', 'CPC $1.84, 日预算$70-130'],
    ['总计', '$30,500-41,500', ''],
]

for i, row_data in enumerate(invest_data, 30):
    write_data_row(ws5, i, row_data)
    if row_data[0] == '总计':
        for col in range(1, 4):
            ws5.cell(row=i, column=col).font = Font(name='Microsoft YaHei', bold=True, size=10, color='1E40AF')

# Go/No-Go Score
ws5.cell(row=38, column=1, value='五、Go/No-Go 五维度评分').font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')

headers_score = ['维度', '权重', '评分(1-10)', '加权分', '说明']
widths_score = [20, 10, 12, 10, 36]
setup_header(ws5, 39, headers_score, widths_score)

score_data = [
    ['市场规模与增长', '25%', 8, 2.00, '类目年增30%；关键词爆发增长40x'],
    ['竞争强度', '25%', 7, 1.75, '$20-50段竞争分散；搜索页无壁垒'],
    ['新品友好度', '20%', 8, 1.60, '新品占比27-31%；低评价即可获销量'],
    ['利润空间', '15%', 6, 0.90, '毛利率~33%尚可；净利率~26-30%（⚠️推测）'],
    ['供应链可行性', '15%', 7, 1.05, '1688资源丰富；需注意认证合规'],
    ['总分', '100%', '', 7.30, '有条件 Go ✅'],
]

for i, row_data in enumerate(score_data, 40):
    write_data_row(ws5, i, row_data, {3, 4})
    if row_data[0] == '总分':
        for col in range(1, 6):
            ws5.cell(row=i, column=col).font = Font(name='Microsoft YaHei', bold=True, size=11, color='1E40AF')
            ws5.cell(row=i, column=col).fill = highlight_fill


# ============================================================
# Save
# ============================================================
output_path = r'C:\Users\Administrator\vacuum_steamer_research\output\20260320_US_vacuum_steamer_市场调研_数据_v1.xlsx'
wb.save(output_path)
print(f'Excel saved to: {output_path}')
print(f'Sheets: {wb.sheetnames}')
