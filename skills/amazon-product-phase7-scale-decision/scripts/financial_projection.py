#!/usr/bin/env python3
"""
规模化决策 - 12 个月财务预测模型
生成乐观/基准/悲观三场景的月度财务预测 Excel。
"""

import json
import sys
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference, LineChart
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def project_12_months(params, scenario_name="baseline"):
    """生成 12 个月预测"""
    base_daily = params.get("daily_sales", 10)
    price = params.get("price", 25)
    cogs = params.get("cogs", 5)
    fba_fee = params.get("fba_fee", 5)
    referral_pct = params.get("referral_pct", 0.15)
    shipping = params.get("shipping_per_unit", 2)
    tacos = params.get("tacos", 0.12)
    return_rate = params.get("return_rate", 0.05)
    storage_per_unit = params.get("storage_per_unit", 0.5)

    # 场景调整
    scenarios_config = {
        "optimistic": {"growth": 0.08, "price_adj": 0, "cogs_adj": -0.05, "tacos_adj": -0.02},
        "baseline": {"growth": 0.03, "price_adj": 0, "cogs_adj": 0, "tacos_adj": 0},
        "pessimistic": {"growth": -0.02, "price_adj": -0.05, "cogs_adj": 0.1, "tacos_adj": 0.05},
    }
    config = scenarios_config.get(scenario_name, scenarios_config["baseline"])

    # 季节性系数（亚马逊美国站）
    seasonality = {
        1: 0.85, 2: 0.80, 3: 0.90, 4: 0.95,
        5: 1.00, 6: 1.05, 7: 1.10, 8: 1.05,
        9: 1.00, 10: 1.15, 11: 1.30, 12: 1.40,
    }

    current_month = datetime.now().month
    months = []

    for i in range(12):
        month_num = ((current_month - 1 + i) % 12) + 1
        month_label = f"M{i + 1}"

        # 销量 = 基础 × (1+增长)^月 × 季节性
        daily = base_daily * ((1 + config["growth"]) ** i) * seasonality[month_num]
        monthly_units = round(daily * 30)

        # 价格调整
        adj_price = price * (1 + config["price_adj"])
        adj_cogs = cogs * (1 + config["cogs_adj"])

        # 成本计算
        revenue = monthly_units * adj_price
        total_cogs = monthly_units * adj_cogs
        total_fba = monthly_units * fba_fee
        total_referral = revenue * referral_pct
        total_shipping = monthly_units * shipping
        total_ad = revenue * (tacos + config["tacos_adj"])
        total_return = revenue * return_rate
        total_storage = monthly_units * storage_per_unit

        total_cost = total_cogs + total_fba + total_referral + total_shipping + total_ad + total_return + total_storage
        net_profit = revenue - total_cost
        net_margin = (net_profit / revenue * 100) if revenue > 0 else 0

        months.append({
            "month_label": month_label,
            "month_num": month_num,
            "daily_sales": round(daily, 1),
            "monthly_units": monthly_units,
            "price": round(adj_price, 2),
            "revenue": round(revenue, 2),
            "cogs": round(total_cogs, 2),
            "fba_fee": round(total_fba, 2),
            "referral_fee": round(total_referral, 2),
            "shipping": round(total_shipping, 2),
            "ad_cost": round(total_ad, 2),
            "return_cost": round(total_return, 2),
            "storage": round(total_storage, 2),
            "total_cost": round(total_cost, 2),
            "net_profit": round(net_profit, 2),
            "net_margin_pct": round(net_margin, 1),
        })

    return months


def write_excel(output_path, params, projections):
    """生成财务预测 Excel"""
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    money_fmt = '$#,##0.00'

    # Sheet 0: 数据来源说明
    ws0 = wb.active
    ws0.title = "数据来源说明"
    ws0.append(["Sheet名", "数据来源", "含义", "日期"])
    ws0.append(["单位经济学", "用户输入 + Sorftime", "当前真实单位成本", datetime.now().strftime("%Y-%m-%d")])
    for scenario in ["乐观", "基准", "悲观"]:
        ws0.append([f"12个月预测_{scenario}", "财务模型计算", f"{scenario}场景月度预测", datetime.now().strftime("%Y-%m-%d")])
    ws0.append(["决策评分卡", "综合评估", "六维加权评分", datetime.now().strftime("%Y-%m-%d")])

    # Sheet 1: 单位经济学
    ws1 = wb.create_sheet("单位经济学")
    ws1.append(["项目", "金额 (USD)", "占售价比"])
    items = [
        ("售价", params.get("price", 0)),
        ("采购成本", -params.get("cogs", 0)),
        ("头程物流", -params.get("shipping_per_unit", 0)),
        ("FBA 配送费", -params.get("fba_fee", 0)),
        ("平台佣金", -params.get("price", 0) * params.get("referral_pct", 0.15)),
        ("广告成本 (TACoS)", -params.get("price", 0) * params.get("tacos", 0.12)),
        ("退货损耗", -params.get("price", 0) * params.get("return_rate", 0.05)),
        ("仓储费", -params.get("storage_per_unit", 0)),
    ]
    total_cost = sum(abs(v) for _, v in items if v < 0)
    net = params.get("price", 0) - total_cost
    for label, val in items:
        ratio = abs(val) / params.get("price", 1) if params.get("price", 0) > 0 else 0
        ws1.append([label, val, ratio])
    ws1.append([])
    ws1.append(["净利润", net, net / params.get("price", 1) if params.get("price", 0) > 0 else 0])

    # Sheets 2-4: 三场景预测
    scenario_names = {"optimistic": "12个月预测_乐观", "baseline": "12个月预测_基准", "pessimistic": "12个月预测_悲观"}
    for scenario_key, sheet_name in scenario_names.items():
        ws = wb.create_sheet(sheet_name)
        months = projections[scenario_key]

        # Header
        ws.append(["月份", "日均销量", "月销量", "售价", "月营收", "采购成本", "FBA费", "佣金", "物流", "广告", "退货", "仓储", "总成本", "净利润", "净利率"])

        for m in months:
            ws.append([
                m["month_label"],
                m["daily_sales"],
                m["monthly_units"],
                m["price"],
                m["revenue"],
                m["cogs"],
                m["fba_fee"],
                m["referral_fee"],
                m["shipping"],
                m["ad_cost"],
                m["return_cost"],
                m["storage"],
                m["total_cost"],
                m["net_profit"],
                m["net_margin_pct"],
            ])

        # 年度汇总
        ws.append([])
        ws.append([
            "年度合计", "",
            sum(m["monthly_units"] for m in months),
            "",
            sum(m["revenue"] for m in months),
            sum(m["cogs"] for m in months),
            sum(m["fba_fee"] for m in months),
            sum(m["referral_fee"] for m in months),
            sum(m["shipping"] for m in months),
            sum(m["ad_cost"] for m in months),
            sum(m["return_cost"] for m in months),
            sum(m["storage"] for m in months),
            sum(m["total_cost"] for m in months),
            sum(m["net_profit"] for m in months),
            round(sum(m["net_profit"] for m in months) / sum(m["revenue"] for m in months) * 100, 1) if sum(m["revenue"] for m in months) > 0 else 0,
        ])

    # Sheet 5: 决策评分卡（模板）
    ws5 = wb.create_sheet("决策评分卡")
    ws5.append(["维度", "权重", "评分 (1-10)", "加权分", "依据"])
    dimensions = [
        ("单位经济学", "25%"),
        ("市场趋势", "20%"),
        ("竞争护城河", "20%"),
        ("供应链就绪", "15%"),
        ("品牌能力", "10%"),
        ("扩展潜力", "10%"),
    ]
    for dim, weight in dimensions:
        ws5.append([dim, weight, "", "", ""])
    ws5.append([])
    ws5.append(["加权总分", "100%", "", "", ""])
    ws5.append(["决策", "", "", "", "Go Big / Maintain / Harvest / Exit"])

    # Format all sheets
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 25)

    wb.save(output_path)
    return output_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description="12 个月财务预测模型")
    parser.add_argument("--input", required=True, help="参数 JSON 路径")
    parser.add_argument("--output", default=None, help="输出 Excel 路径")
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        params = json.load(f)

    projections = {}
    for scenario in ["optimistic", "baseline", "pessimistic"]:
        projections[scenario] = project_12_months(params, scenario)

    output_path = args.output
    if not output_path:
        base = os.path.dirname(args.input)
        date_str = datetime.now().strftime("%Y%m%d")
        asin = params.get("asin", "product")
        site = params.get("site", "US")
        output_path = os.path.join(base, f"{date_str}_{site}_{asin}_财务预测模型.xlsx")

    write_excel(output_path, params, projections)

    # Print summary
    baseline = projections["baseline"]
    annual_revenue = sum(m["revenue"] for m in baseline)
    annual_profit = sum(m["net_profit"] for m in baseline)

    print("=== 12 个月财务预测 ===")
    print(f"基准场景:")
    print(f"  年营收: ${annual_revenue:,.2f}")
    print(f"  年利润: ${annual_profit:,.2f}")
    print(f"  年利润率: {annual_profit / annual_revenue * 100:.1f}%" if annual_revenue > 0 else "  年利润率: N/A")
    print(f"\n乐观场景年利润: ${sum(m['net_profit'] for m in projections['optimistic']):,.2f}")
    print(f"悲观场景年利润: ${sum(m['net_profit'] for m in projections['pessimistic']):,.2f}")
    print(f"\nExcel 已保存: {output_path}")
    print("projection_ok")


if __name__ == "__main__":
    main()
