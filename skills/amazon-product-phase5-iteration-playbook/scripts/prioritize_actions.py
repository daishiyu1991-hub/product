#!/usr/bin/env python3
"""
迭代优化 - 优先级排序器
基于 Impact x Effort 评分生成优先级排序。
"""

import json
import sys
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def classify_quadrant(impact, effort):
    """分类到四象限"""
    if impact >= 3 and effort <= 3:
        return "⭐ Quick Win"
    elif impact >= 3 and effort > 3:
        return "💎 High Value"
    elif impact < 3 and effort <= 3:
        return "📋 Fill In"
    else:
        return "❌ Avoid"


def prioritize(actions):
    """计算优先级并排序"""
    for action in actions:
        impact = action.get("impact", 1)
        effort = action.get("effort", 5)
        priority = round(impact / effort, 2) if effort > 0 else 0
        action["priority"] = priority
        action["quadrant"] = classify_quadrant(impact, effort)

    # 按优先级降序排序
    actions.sort(key=lambda x: x["priority"], reverse=True)
    return actions


def write_excel(output_path, actions):
    """生成优先级矩阵 Excel"""
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    qw_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    hv_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fi_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    av_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    quadrant_fills = {
        "⭐ Quick Win": qw_fill,
        "💎 High Value": hv_fill,
        "📋 Fill In": fi_fill,
        "❌ Avoid": av_fill,
    }

    # Sheet 1: 优先级排序
    ws1 = wb.active
    ws1.title = "优先级排序"
    ws1.append(["排名", "优化项", "类别", "Impact (1-5)", "Effort (1-5)", "Priority", "象限", "预期效果", "数据依据"])

    for i, action in enumerate(actions, 1):
        row_num = ws1.max_row + 1
        ws1.append([
            i,
            action.get("name", ""),
            action.get("category", ""),
            action.get("impact", ""),
            action.get("effort", ""),
            action.get("priority", ""),
            action.get("quadrant", ""),
            action.get("expected_effect", ""),
            action.get("data_basis", ""),
        ])
        fill = quadrant_fills.get(action.get("quadrant", ""), None)
        if fill:
            ws1.cell(row=row_num, column=7).fill = fill

    # Sheet 2: 四象限汇总
    ws2 = wb.create_sheet("四象限汇总")
    for quadrant_name in ["⭐ Quick Win", "💎 High Value", "📋 Fill In", "❌ Avoid"]:
        ws2.append([quadrant_name])
        ws2.append(["优化项", "Impact", "Effort", "Priority", "预期效果"])
        items = [a for a in actions if a.get("quadrant") == quadrant_name]
        for item in items:
            ws2.append([
                item.get("name", ""),
                item.get("impact", ""),
                item.get("effort", ""),
                item.get("priority", ""),
                item.get("expected_effect", ""),
            ])
        ws2.append([])

    # Format headers
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill

    for col in ws1.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

    wb.save(output_path)
    return output_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description="迭代优化 - 优先级排序器")
    parser.add_argument("--input", required=True, help="行动项 JSON 路径")
    parser.add_argument("--output", default=None, help="输出 Excel 路径")
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)

    actions = data.get("actions", data) if isinstance(data, dict) else data
    prioritized = prioritize(actions)

    output_path = args.output or os.path.join(os.path.dirname(args.input), "iteration_priority.xlsx")
    write_excel(output_path, prioritized)

    print("=== 迭代优化优先级排序 ===")
    print(f"总计 {len(prioritized)} 个优化项")
    for q in ["⭐ Quick Win", "💎 High Value", "📋 Fill In", "❌ Avoid"]:
        count = sum(1 for a in prioritized if a.get("quadrant") == q)
        print(f"  {q}: {count} 个")
    print(f"\nTop 3 优先项:")
    for i, a in enumerate(prioritized[:3], 1):
        print(f"  {i}. {a.get('name', '')} (Priority: {a.get('priority', '')}, {a.get('quadrant', '')})")
    print(f"\nExcel 已保存: {output_path}")
    print("prioritize_ok")


if __name__ == "__main__":
    main()
