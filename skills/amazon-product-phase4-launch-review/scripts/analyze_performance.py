#!/usr/bin/env python3
"""
上架复盘 - 性能分析器
解析 Sorftime 产品数据，计算 KPI 对照表。
"""

import json
import sys
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# Kill/Continue/Pivot 阈值
DEFAULT_THRESHOLDS = {
    "conversion_rate": {"green": 10, "yellow": 5, "red": 5},  # ≥green=🟢, ≥yellow=🟡, <red=🔴
    "bsr_trend": {"green": "stable_or_up", "yellow": "volatile", "red": "declining"},
    "acos": {"green": 30, "yellow": 50, "red": 50},  # ≤green=🟢, ≤yellow=🟡, >red=🔴
    "return_rate": {"green": 8, "yellow": 15, "red": 15},
    "rating": {"green": 4.0, "yellow": 3.5, "red": 3.5},
    "organic_share": {"green": 40, "yellow": 20, "red": 20},
}


def evaluate_traffic_light(metric_name, value, thresholds=None):
    """评估单个指标的红绿灯状态"""
    if thresholds is None:
        thresholds = DEFAULT_THRESHOLDS.get(metric_name, {})

    if not thresholds or value is None:
        return "⚪", "未验证"

    # 越大越好的指标
    if metric_name in ("conversion_rate", "rating", "organic_share"):
        if value >= thresholds.get("green", 0):
            return "🟢", "Continue"
        elif value >= thresholds.get("yellow", 0):
            return "🟡", "Pivot"
        else:
            return "🔴", "Kill"

    # 越小越好的指标
    elif metric_name in ("acos", "return_rate"):
        if value <= thresholds.get("green", 100):
            return "🟢", "Continue"
        elif value <= thresholds.get("yellow", 100):
            return "🟡", "Pivot"
        else:
            return "🔴", "Kill"

    return "⚪", "未评估"


def analyze_performance(data, baseline=None):
    """分析产品表现"""
    results = {}

    # 从数据中提取指标
    metrics = data.get("metrics", {})

    indicators = [
        ("conversion_rate", "转化率 (%)", metrics.get("conversion_rate")),
        ("acos", "ACOS (%)", metrics.get("acos")),
        ("return_rate", "退货率 (%)", metrics.get("return_rate")),
        ("rating", "评分", metrics.get("rating")),
        ("organic_share", "自然流量占比 (%)", metrics.get("organic_share")),
    ]

    custom_thresholds = baseline.get("test_criteria", {}) if baseline else {}

    traffic_lights = []
    for metric_id, label, value in indicators:
        thresholds = custom_thresholds.get(metric_id, DEFAULT_THRESHOLDS.get(metric_id, {}))
        light, decision = evaluate_traffic_light(metric_id, value, thresholds)

        baseline_pass = None
        baseline_fail = None
        if baseline and "test_criteria" in baseline:
            for tc in baseline.get("test_criteria_list", []):
                if tc.get("metric_id") == metric_id:
                    baseline_pass = tc.get("pass_threshold")
                    baseline_fail = tc.get("fail_threshold")

        traffic_lights.append({
            "metric_id": metric_id,
            "label": label,
            "current_value": value,
            "baseline_pass": baseline_pass,
            "baseline_fail": baseline_fail,
            "light": light,
            "decision": decision,
            "trend": metrics.get(f"{metric_id}_trend", "→"),
        })

    # BSR 趋势单独处理
    bsr_trend = metrics.get("bsr_trend", "unknown")
    bsr_light = "🟢" if bsr_trend in ("stable", "up") else "🟡" if bsr_trend == "volatile" else "🔴"
    bsr_decision = "Continue" if bsr_light == "🟢" else "Pivot" if bsr_light == "🟡" else "Kill"
    traffic_lights.append({
        "metric_id": "bsr_trend",
        "label": "BSR 趋势",
        "current_value": bsr_trend,
        "light": bsr_light,
        "decision": bsr_decision,
        "trend": "—",
    })

    # 最终决策
    greens = sum(1 for t in traffic_lights if t["light"] == "🟢")
    yellows = sum(1 for t in traffic_lights if t["light"] == "🟡")
    reds = sum(1 for t in traffic_lights if t["light"] == "🔴")

    # 一票否决
    veto_metrics = ["conversion_rate", "rating"]
    veto_kill = any(
        t["light"] == "🔴" for t in traffic_lights if t["metric_id"] in veto_metrics
    )
    veto_count = sum(1 for t in traffic_lights if t["metric_id"] in veto_metrics and t["light"] == "🔴")

    if veto_count >= 2:
        final_decision = "Kill"
        final_reason = "转化率+评分双红灯，一票否决"
    elif reds >= 3:
        final_decision = "Kill"
        final_reason = f"红灯 {reds} 个 ≥ 3"
    elif yellows >= 3:
        final_decision = "Pivot"
        final_reason = f"黄灯 {yellows} 个 ≥ 3"
    elif greens >= 4 and reds == 0:
        final_decision = "Continue"
        final_reason = f"绿灯 {greens} 个 ≥ 4，无红灯"
    elif greens >= 3 and reds <= 1:
        final_decision = "Conditional Continue"
        final_reason = f"绿灯 {greens} 个，红灯 {reds} 个 ≤ 1，需重点解决红灯指标"
    else:
        final_decision = "Pivot"
        final_reason = f"绿灯 {greens}，黄灯 {yellows}，红灯 {reds}"

    results["traffic_lights"] = traffic_lights
    results["summary"] = {
        "greens": greens,
        "yellows": yellows,
        "reds": reds,
        "final_decision": final_decision,
        "reason": final_reason,
        "veto_triggered": veto_kill,
    }

    return results


def write_excel(output_path, data, results):
    """生成复盘数据 Excel"""
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Sheet 1: 数据来源说明
    ws1 = wb.active
    ws1.title = "数据来源说明"
    ws1.append(["Sheet名", "数据来源", "含义", "日期"])
    ws1.append(["KPI对照表", "Sorftime + 用户输入", "核心指标 vs MVP 基线", datetime.now().strftime("%Y-%m-%d")])
    ws1.append(["红绿灯评估", "计算", "Kill/Continue/Pivot 决策", datetime.now().strftime("%Y-%m-%d")])

    # Sheet 2: KPI 对照表
    ws2 = wb.create_sheet("KPI对照表")
    ws2.append(["指标", "当前值", "MVP Pass 阈值", "MVP Fail 阈值", "状态", "趋势"])
    for tl in results["traffic_lights"]:
        row_num = ws2.max_row + 1
        ws2.append([
            tl["label"],
            tl.get("current_value", ""),
            tl.get("baseline_pass", ""),
            tl.get("baseline_fail", ""),
            f"{tl['light']} {tl['decision']}",
            tl.get("trend", ""),
        ])
        # Color the status cell
        light = tl["light"]
        fill = green_fill if light == "🟢" else yellow_fill if light == "🟡" else red_fill if light == "🔴" else None
        if fill:
            ws2.cell(row=row_num, column=5).fill = fill

    # Sheet 3: 红绿灯评估
    ws3 = wb.create_sheet("红绿灯评估")
    summary = results["summary"]
    ws3.append(["项目", "值"])
    ws3.append(["🟢 绿灯数", summary["greens"]])
    ws3.append(["🟡 黄灯数", summary["yellows"]])
    ws3.append(["🔴 红灯数", summary["reds"]])
    ws3.append(["一票否决触发", "是" if summary["veto_triggered"] else "否"])
    ws3.append(["最终决策", summary["final_decision"]])
    ws3.append(["决策理由", summary["reason"]])

    # Format headers
    for ws in [ws1, ws2, ws3]:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

    wb.save(output_path)
    return output_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description="上架复盘 - 性能分析器")
    parser.add_argument("--input", required=True, help="产品数据 JSON 路径")
    parser.add_argument("--baseline", default=None, help="MVP 蓝图参数 JSON (可选)")
    parser.add_argument("--output", default=None, help="输出 Excel 路径")
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)

    baseline = None
    if args.baseline and os.path.exists(args.baseline):
        with open(args.baseline, "r", encoding="utf-8") as f:
            baseline = json.load(f)

    results = analyze_performance(data, baseline)

    output_path = args.output
    if not output_path:
        base = os.path.dirname(args.input)
        output_path = os.path.join(base, "review_analysis.xlsx")

    write_excel(output_path, data, results)

    summary = results["summary"]
    print("=== 上架复盘分析结果 ===")
    print(f"🟢 绿灯: {summary['greens']}")
    print(f"🟡 黄灯: {summary['yellows']}")
    print(f"🔴 红灯: {summary['reds']}")
    print(f"决策: {summary['final_decision']}")
    print(f"理由: {summary['reason']}")
    print(f"Excel 已保存: {output_path}")
    print("analyze_ok")


if __name__ == "__main__":
    main()
