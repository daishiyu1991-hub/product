#!/usr/bin/env python3
"""
MVP 蓝图 - 盈亏平衡计算器
计算单位成本、毛利率、盈亏平衡点，输出 Excel 财务模型。
"""

import json
import sys
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def calculate_unit_economics(params):
    """计算单位经济学"""
    price = params.get("price", 0)
    cogs = params.get("cogs", 0)  # 采购成本
    shipping_to_fba = params.get("shipping_to_fba", 0)  # 头程物流/件
    fba_fee = params.get("fba_fee", 0)  # FBA 配送费
    referral_pct = params.get("referral_pct", 0.15)  # 平台佣金比例
    ad_cost_per_unit = params.get("ad_cost_per_unit", 0)  # 广告成本/件
    return_rate = params.get("return_rate", 0.05)  # 退货率
    return_damage_rate = params.get("return_damage_rate", 0.5)  # 退货损毁率
    storage_per_unit = params.get("storage_per_unit", 0)  # 仓储费/件/月

    referral_fee = price * referral_pct
    return_cost = price * return_rate * return_damage_rate
    variable_cost = cogs + shipping_to_fba + fba_fee + referral_fee + ad_cost_per_unit + return_cost + storage_per_unit
    gross_profit = price - variable_cost
    gross_margin = (gross_profit / price * 100) if price > 0 else 0

    return {
        "price": price,
        "cogs": cogs,
        "shipping_to_fba": shipping_to_fba,
        "fba_fee": fba_fee,
        "referral_fee": referral_fee,
        "ad_cost_per_unit": ad_cost_per_unit,
        "return_cost": return_cost,
        "storage_per_unit": storage_per_unit,
        "variable_cost_total": variable_cost,
        "gross_profit": gross_profit,
        "gross_margin_pct": gross_margin,
    }


def calculate_breakeven(params, unit_economics):
    """计算盈亏平衡"""
    fixed_costs = {
        "mold_cost": params.get("mold_cost", 0),
        "packaging_design": params.get("packaging_design", 0),
        "certification": params.get("certification", 0),
        "photography": params.get("photography", 0),
        "initial_ad_budget": params.get("initial_ad_budget", 0),
        "other_fixed": params.get("other_fixed", 0),
    }
    total_fixed = sum(fixed_costs.values())
    contribution = unit_economics["gross_profit"]
    breakeven_units = int(total_fixed / contribution) + 1 if contribution > 0 else float("inf")
    target_daily = params.get("target_daily_sales", 10)
    breakeven_days = breakeven_units / target_daily if target_daily > 0 else float("inf")

    first_batch = params.get("first_batch_qty", 500)
    first_batch_cost = first_batch * params.get("cogs", 0)
    total_investment = total_fixed + first_batch_cost + first_batch * params.get("shipping_to_fba", 0)

    return {
        "fixed_costs": fixed_costs,
        "total_fixed_costs": total_fixed,
        "breakeven_units": breakeven_units,
        "breakeven_days": round(breakeven_days, 1),
        "target_daily_sales": target_daily,
        "first_batch_qty": first_batch,
        "first_batch_cost": first_batch_cost,
        "total_investment": total_investment,
    }


def generate_scenarios(params, unit_economics):
    """生成乐观/基准/悲观三场景"""
    price = params.get("price", 0)
    daily = params.get("target_daily_sales", 10)

    scenarios = {}
    for name, multipliers in [
        ("optimistic", {"sales": 1.5, "price": 1.0, "cogs": 0.95, "acos_adj": 0.8}),
        ("baseline", {"sales": 1.0, "price": 1.0, "cogs": 1.0, "acos_adj": 1.0}),
        ("pessimistic", {"sales": 0.5, "price": 0.9, "cogs": 1.1, "acos_adj": 1.3}),
    ]:
        adj_daily = daily * multipliers["sales"]
        adj_price = price * multipliers["price"]
        adj_cogs = params.get("cogs", 0) * multipliers["cogs"]
        adj_ad = params.get("ad_cost_per_unit", 0) * multipliers["acos_adj"]

        adj_referral = adj_price * params.get("referral_pct", 0.15)
        adj_return = adj_price * params.get("return_rate", 0.05) * params.get("return_damage_rate", 0.5)
        adj_variable = adj_cogs + params.get("shipping_to_fba", 0) + params.get("fba_fee", 0) + adj_referral + adj_ad + adj_return
        adj_profit = adj_price - adj_variable
        monthly_revenue = adj_daily * 30 * adj_price
        monthly_profit = adj_daily * 30 * adj_profit

        scenarios[name] = {
            "daily_sales": round(adj_daily, 1),
            "price": round(adj_price, 2),
            "unit_profit": round(adj_profit, 2),
            "monthly_revenue": round(monthly_revenue, 2),
            "monthly_profit": round(monthly_profit, 2),
            "margin_pct": round((adj_profit / adj_price * 100) if adj_price > 0 else 0, 1),
        }
    return scenarios


def write_excel(output_path, params, unit_economics, breakeven, scenarios):
    """生成 Excel 文件"""
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    money_fmt = '$#,##0.00'
    pct_fmt = '0.0%'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Sheet 1: 数据来源说明
    ws1 = wb.active
    ws1.title = "数据来源说明"
    ws1.append(["Sheet名", "数据来源", "数据含义", "关键参数", "计算日期"])
    ws1.append(["产品成本明细", "用户输入 / Sorftime ali1688", "单位成本拆解", "—", datetime.now().strftime("%Y-%m-%d")])
    ws1.append(["盈亏平衡计算", "计算", "固定成本+盈亏平衡点", "—", datetime.now().strftime("%Y-%m-%d")])
    ws1.append(["多场景模拟", "计算", "乐观/基准/悲观预测", "—", datetime.now().strftime("%Y-%m-%d")])
    for cell in ws1[1]:
        cell.font = header_font_white
        cell.fill = header_fill

    # Sheet 2: 产品成本明细
    ws2 = wb.create_sheet("产品成本明细")
    cost_items = [
        ("售价", unit_economics["price"]),
        ("采购成本", -unit_economics["cogs"]),
        ("头程物流", -unit_economics["shipping_to_fba"]),
        ("FBA 配送费", -unit_economics["fba_fee"]),
        ("平台佣金 (15%)", -unit_economics["referral_fee"]),
        ("广告成本/件", -unit_economics["ad_cost_per_unit"]),
        ("退货损耗", -unit_economics["return_cost"]),
        ("仓储费/件", -unit_economics["storage_per_unit"]),
        ("", ""),
        ("单位总成本", -unit_economics["variable_cost_total"]),
        ("单位毛利", unit_economics["gross_profit"]),
        ("毛利率", unit_economics["gross_margin_pct"] / 100),
    ]
    ws2.append(["项目", "金额 (USD)", "占售价比"])
    for item, val in cost_items:
        if item == "":
            ws2.append([])
            continue
        if item == "毛利率":
            ws2.append([item, val, ""])
        else:
            ratio = abs(val) / unit_economics["price"] if unit_economics["price"] > 0 else 0
            ws2.append([item, val, ratio])

    # Sheet 3: 盈亏平衡计算
    ws3 = wb.create_sheet("盈亏平衡计算")
    ws3.append(["固定成本项", "金额 (USD)"])
    for k, v in breakeven["fixed_costs"].items():
        label_map = {
            "mold_cost": "模具/开模费",
            "packaging_design": "包装设计",
            "certification": "认证费用",
            "photography": "产品拍摄",
            "initial_ad_budget": "首月广告预算",
            "other_fixed": "其他固定费用",
        }
        ws3.append([label_map.get(k, k), v])
    ws3.append(["固定成本合计", breakeven["total_fixed_costs"]])
    ws3.append([])
    ws3.append(["盈亏平衡指标", "值"])
    ws3.append(["单位贡献利润", unit_economics["gross_profit"]])
    ws3.append(["盈亏平衡销量 (件)", breakeven["breakeven_units"]])
    ws3.append(["目标日均销量", breakeven["target_daily_sales"]])
    ws3.append(["盈亏平衡天数", breakeven["breakeven_days"]])
    ws3.append([])
    ws3.append(["首批投入", "值"])
    ws3.append(["首批采购量", breakeven["first_batch_qty"]])
    ws3.append(["首批采购成本", breakeven["first_batch_cost"]])
    ws3.append(["总投入 (含固定+采购+物流)", breakeven["total_investment"]])

    # Sheet 4: 多场景模拟
    ws4 = wb.create_sheet("多场景模拟")
    ws4.append(["指标", "乐观", "基准", "悲观"])
    metrics = ["daily_sales", "price", "unit_profit", "monthly_revenue", "monthly_profit", "margin_pct"]
    labels = ["日均销量 (件)", "售价 (USD)", "单位利润 (USD)", "月营收 (USD)", "月利润 (USD)", "利润率 (%)"]
    for label, metric in zip(labels, metrics):
        ws4.append([
            label,
            scenarios["optimistic"][metric],
            scenarios["baseline"][metric],
            scenarios["pessimistic"][metric],
        ])

    # Format headers for all sheets
    for ws in [ws1, ws2, ws3, ws4]:
        for cell in ws[1]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

    wb.save(output_path)
    return output_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description="MVP 盈亏平衡计算器")
    parser.add_argument("--input", required=True, help="参数 JSON 文件路径")
    parser.add_argument("--output", default=None, help="输出 Excel 路径 (默认同目录)")
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        params = json.load(f)

    unit_economics = calculate_unit_economics(params)
    breakeven = calculate_breakeven(params, unit_economics)
    scenarios = generate_scenarios(params, unit_economics)

    output_path = args.output
    if not output_path:
        base = os.path.dirname(args.input)
        date_str = datetime.now().strftime("%Y%m%d")
        site = params.get("site", "US")
        keyword = params.get("keyword", "product")
        output_path = os.path.join(base, f"{date_str}_{site}_{keyword}_MVP蓝图_财务模型.xlsx")

    write_excel(output_path, params, unit_economics, breakeven, scenarios)

    # Print summary
    print(f"=== MVP 盈亏平衡计算结果 ===")
    print(f"售价: ${unit_economics['price']:.2f}")
    print(f"单位毛利: ${unit_economics['gross_profit']:.2f}")
    print(f"毛利率: {unit_economics['gross_margin_pct']:.1f}%")
    print(f"盈亏平衡销量: {breakeven['breakeven_units']} 件")
    print(f"盈亏平衡天数: {breakeven['breakeven_days']} 天")
    print(f"总投入: ${breakeven['total_investment']:.2f}")
    print(f"Excel 已保存: {output_path}")
    print("calculate_ok")


if __name__ == "__main__":
    main()
